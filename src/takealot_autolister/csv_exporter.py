"""
Takealot Loadsheet 生成器

根据 ListingDraft + ProductSource，找到对应类目的 loadsheet 模板，
填入产品数据，生成可直接上传到 Takealot 的 .xlsm 文件。

用法：
    from takealot_autolister.csv_exporter import generate_loadsheet
    out_path = generate_loadsheet(draft, source, run_dir)
"""
from __future__ import annotations

import json
import re
import shutil
from pathlib import Path
from typing import Any

import openpyxl

from .types import ListingDraft, ProductSource

# ─── 路径 ──────────────────────────────────────────────────────────────────────
_HERE = Path(__file__).resolve().parent

# 一些字段（尤其是颜色）里会填类似 "Not specified" 这种占位符，生成 loadsheet 时应视为未填
_PLACEHOLDER_STRS = {
    "to be confirmed",
    "tbc",
    "n/a",
    "unknown",
    "please confirm",
    "not specified",
}


def _non_placeholder(v: Any) -> str:
    """把占位符值（To be confirmed 等）转为空字符串，让 AI 补填步骤处理。"""
    s = str(v or "").strip()
    return "" if s.lower() in _PLACEHOLDER_STRS else s
_ROOT = _HERE.parent.parent
_RAW_DIR = _ROOT / "input" / "loadsheets" / "raw"

# ─── 中文类目 → 英文关键词映射（1688 中文类目 → Takealot 英文类目搜索词）──────
_ZH_TO_EN: dict[str, str] = {
    # ── 耳机 ──
    "蓝牙耳机": "Cellphone Headsets",
    "耳机": "Cellphone Headsets",
    "耳麦": "Cellphone Headsets",
    "入耳式耳机": "Cellphone Headsets",
    "头戴式耳机": "Standard Headphones",
    "游戏耳机": "Gaming Headphones",
    "DJ耳机": "DJ Headphones",
    "运动耳机": "Sport Headphones",
    # ── 音箱 ──
    "蓝牙音箱": "Speakers",
    "音箱": "Speakers",
    "音响": "Speakers",
    "无线音箱": "Speakers",
    "便携音箱": "Speakers",
    "蓝牙喇叭": "Speakers",
    "智能音箱": "Smart Audio",
    "回音壁": "Soundbars",
    "条形音箱": "Soundbars",
    "低音炮": "Subwoofers",
    "功放": "Audio Amplifiers",
    "麦克风": "Microphones",
    "话筒": "Microphones",
    # ── 手机 & 平板 ──
    "手机": "Mobile Phones",
    "智能手机": "Mobile Phones",
    "平板": "Tablets and E-Readers",
    "平板电脑": "Tablets and E-Readers",
    "电子书": "Tablets and E-Readers",
    # ── 手机配件 ──
    "手机壳": "Mobile Phone Cases",
    "手机套": "Mobile Phone Cases",
    "手机支架": "Mobile Phone Mounts & Stands",
    "手机膜": "Screen Protectors",
    "钢化膜": "Screen Protectors",
    "手机镜头": "Mobile Phone Camera Accessories",
    "自拍杆": "Mobile Phone Camera Accessories",
    "手机散热器": "Mobile Phone Accessories",
    # ── 充电 ──
    "充电宝": "Power Banks",
    "移动电源": "Power Banks",
    "充电器": "Power Adapters & Chargers",
    "无线充电器": "Power Adapters & Chargers",
    "数据线": "Cellphone Cables",
    "充电线": "Cellphone Cables",
    "type-c线": "Cellphone Cables",
    # ── 存储 ──
    "内存卡": "Memory Cards",
    "TF卡": "Memory Cards",
    "SD卡": "Memory Cards",
    "读卡器": "Memory Card Readers",
    "U盘": "USB Flash Drives",
    "硬盘": "Portable HDD",
    "移动硬盘": "Portable HDD",
    "固态硬盘": "Portable SSD",
    "移动固态": "Portable SSD",
    "PSSD": "Portable SSD",
    "SSD": "Portable SSD",
    # ── 电脑配件 ──
    "鼠标": "Input Devices",
    "键盘": "Keyboards",
    "鼠标垫": "Mouse Pads",
    "机械键盘": "Keyboards",
    "游戏键盘": "Keyboards",
    "游戏鼠标": "Input Devices",
    "显示器": "Computer Monitors",
    "电脑包": "Bags & Cases",
    "笔记本包": "Laptop Bags",
    "散热器": "Computer System Cooling Parts",
    "网卡": "Networking",
    "路由器": "Networking",
    "USB集线器": "USB & FireWire Hubs",
    # ── 电脑整机 ──
    "笔记本": "Computers",
    "台式机": "Computers",
    "电脑主机": "Computers",
    # ── 摄影 ──
    "相机": "Cameras",
    "运动相机": "Action Cameras",
    "无人机": "Drones",
    "三脚架": "Tripods & Monopods",
    "摄像头": "Cameras",
    "相机包": "Camera Parts & Accessories",
    # ── 智能穿戴 ──
    # 注："Wearable Tech" 被 loadsheet-101 占用，需映射到 loadsheet-258 的精确子类
    "智能手表": "Smart Watches",       # -> loadsheet-258
    "手环": "Activity Trackers",       # -> loadsheet-258
    "智能手环": "Activity Trackers",   # -> loadsheet-258
    "运动手表": "Sport Watches",       # -> loadsheet-258
    "智能戒指": "Activity Trackers",   # -> loadsheet-258 (Digital Rings 也在258)
    "智能指环": "Activity Trackers",   # -> loadsheet-258
    "健康戒指": "Activity Trackers",   # -> loadsheet-258
    "运动追踪器": "Activity Trackers", # -> loadsheet-258
    "活动追踪器": "Activity Trackers", # -> loadsheet-258
    "健身追踪器": "Activity Trackers", # -> loadsheet-258
    "计步器": "Activity Trackers",     # -> loadsheet-258
    # ── 游戏 ──
    "游戏手柄": "Video Game Accessories",
    "游戏机": "Video Game Consoles",
    "游戏耳机": "Gaming Headphones",
    "游戏椅": "Gaming Chairs",
    # ── 电池 & 电源 ──
    "电池": "Batteries",
    "干电池": "Batteries",
    "充电电池": "Batteries",
    "发电机": "Generators",
    # ── 智能家居 ──
    "智能灯": "Smart Lighting",
    "智能插座": "Smart Energy Solutions",
    "安防摄像头": "Business & Home Security",
    "门铃": "Intercoms",
    # ── 汽车配件 ──
    "车载充电器": "Motor Vehicle Electronics",
    "车载支架": "Motor Vehicle Electronics",
    "行车记录仪": "Motor Vehicle Electronics",
    "汽车音响": "Motor Vehicle Electronics",
    "无线CarPlay": "Motor Vehicle Electronics",
    "有线CarPlay": "Motor Vehicle Electronics",
    "CarPlay": "Motor Vehicle Electronics",
    "车机": "Motor Vehicle Electronics",
    "车载导航": "Motor Vehicle Electronics",
    "倒车摄像头": "Motor Vehicle Electronics",
    "停车摄像头": "Motor Vehicle Electronics",
    "车载蓝牙": "Motor Vehicle Electronics",
    # ── 家用电器 ──
    "电风扇": "Household Appliances",
    "空气净化器": "Household Appliances",
    "加湿器": "Household Appliances",
    "吸尘器": "Household Appliances",
    "电热水壶": "Kitchen Appliances",
    "咖啡机": "Kitchen Appliances",
    "榨汁机": "Kitchen Appliances",
    # ── 健康 ──
    "血压计": "Health Care",
    "体温计": "Health Care",
    "按摩仪": "Health Care",
    # ── 运动户外 ──
    "跑步机": "Cardio",
    "哑铃": "Strength Training",
    "瑜伽垫": "Yoga & Pilates",
    "帐篷": "Camping & Hiking",
    "背包": "Backpacks",
    # ── 办公 ──
    "打印机": "Print, Copy, Scan & Fax",
    "扫描仪": "Print, Copy, Scan & Fax",
    "投影仪": "Projectors",
    "台灯": "Lighting",
    "文具": "General Office Supplies",
    # ── 玩具 ──
    "玩具": "Toys",
    "积木": "Toys",
    "遥控车": "Toys",
    "桌游": "Board Games",
    # ── 美容 ──
    "电动牙刷": "Oral Care",
    "吹风机": "Hair Dryers",
    "电吹风": "Hair Dryers",
    "卷发棒": "Curlers & Sets",
    "卷发器": "Curlers & Sets",
    "烫发器": "Curlers & Sets",
    "直发器": "Irons & Straighteners",
    "直发板": "Irons & Straighteners",
    "直发夹": "Irons & Straighteners",
    "拉直板": "Irons & Straighteners",
    "护肤品": "Skin Care",
    "口红": "Makeup",
    # ── 宠物 ──
    "狗粮": "Dog Supplies",
    "猫粮": "Cat Supplies",
    "宠物用品": "Dog Supplies",
    # ── 通用 ──
    "适配器": "Adapters",
    "转接头": "Adapters",
    "电缆": "Cables",
    "线缆": "Cables",
}


def _translate_zh_category(parts: list[str]) -> list[str]:
    """把中文类目列表翻译为英文，方便 loadsheet 匹配。"""
    result = []
    for p in parts:
        p_strip = p.strip()
        # 直接命中
        if p_strip in _ZH_TO_EN:
            result.append(_ZH_TO_EN[p_strip])
            continue
        # 部分匹配
        matched = next((v for k, v in _ZH_TO_EN.items() if k in p_strip or p_strip in k), None)
        if matched:
            result.append(matched)
        else:
            result.append(p_strip)  # 保留原文，让模糊匹配再试
    return result



# ─── 类目 → loadsheet 映射（启动时从 raw xlsm 建立，缓存到模块变量）──────────
_CATEGORY_MAP: dict[str, str] | None = None   # lowercase_category_name -> xlsm 路径


def _build_category_map() -> dict[str, str]:
    """扫描所有 raw xlsm，建立 lowest_category_name → xlsm_path 映射。"""
    mapping: dict[str, str] = {}
    for xlsm in sorted(_RAW_DIR.glob("*.xlsm")):
        try:
            wb = openpyxl.load_workbook(str(xlsm), read_only=True, keep_vba=False)
            if "Category Tree Lookup" not in wb.sheetnames:
                wb.close()
                continue
            ws = wb["Category Tree Lookup"]
            for row in ws.iter_rows(min_row=4, values_only=True):
                if not row or not row[0]:
                    continue
                main = str(row[0]).strip()
                low  = str(row[1]).strip() if len(row) > 1 and row[1] else main
                # 用 main category 做主键（不含 ID 的纯名称也建一条）
                key_main  = _strip_id(main).lower()
                key_low   = _strip_id(low).lower()
                # 末级路径（去掉多级前缀，如 "A->B->C" 取 "C"）
                key_leaf  = key_low.split("->")[-1].strip()
                for k in {key_main, key_low, key_leaf}:
                    if k and k not in mapping:
                        mapping[k] = str(xlsm)
            wb.close()
        except Exception:
            pass
    return mapping


def _get_category_map() -> dict[str, str]:
    global _CATEGORY_MAP
    if _CATEGORY_MAP is None:
        _CATEGORY_MAP = _build_category_map()
    return _CATEGORY_MAP


def _strip_id(text: str) -> str:
    """去掉括号内的数字 ID，如 'Cellphone Headsets (21654)' → 'Cellphone Headsets'。"""
    return re.sub(r"\s*\(\d+\)\s*$", "", text).strip()


def _norm(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip().lower()


# ─── LLM 自动映射类目 ─────────────────────────────────────────────────────────

def _llm_map_category(category_path: list[str], draft: "ListingDraft") -> str | None:
    """
    当规则映射找不到 loadsheet 时，用 LLM 从现有类目列表中选最匹配的。
    返回 xlsm 路径，失败返回 None。
    """
    cmap = _get_category_map()
    if not cmap:
        return None

    # 所有可选类目（去重，最多取 120 个关键 key）
    all_keys = sorted({k for k in cmap if len(k) > 3})[:120]
    cats_str = "\n".join(all_keys)

    product_desc = f"Product: {draft.title}\nSupplier category: {' > '.join(category_path)}"
    if draft.attributes:
        attrs_sample = ", ".join(f"{k}:{v}" for k, v in list(draft.attributes.items())[:10] if v)
        product_desc += f"\nAttributes: {attrs_sample}"

    prompt = (
        f"{product_desc}\n\n"
        "Below are available Takealot category names.\n"
        "Choose the SINGLE best matching category for this product.\n"
        "Return JSON only: {\"category\": \"<exact category name from the list>\"}\n\n"
        f"Available categories:\n{cats_str}"
    )

    try:
        from .llm import _call_llm_json, is_llm_available
        if not is_llm_available():
            return None
        result = _call_llm_json(prompt, temperature=0.1)
        chosen = str(result.get("category", "")).strip().lower()
        if chosen and chosen in cmap:
            return cmap[chosen]
        # 模糊匹配
        for k, v in cmap.items():
            if chosen and (chosen in k or k in chosen):
                return v
    except Exception as e:
        print(f"[csv_exporter] LLM 类目映射出错：{e}")
    return None


# ─── 查找 loadsheet ────────────────────────────────────────────────────────────

def find_loadsheet(category_path: list[str]) -> str | None:
    """
    给定类目路径（如 ['Consumer Electronics', 'Electronic Accessories', 'Cellphone Headsets']），
    返回对应的 xlsm 路径；找不到返回 None。
    """
    cmap = _get_category_map()
    # 从最末级向上逐级尝试
    for part in reversed(category_path):
        part_n = _norm(_strip_id(part))
        if part_n in cmap:
            return cmap[part_n]
        # 模糊匹配：任一候选 key 包含 part_n
        for k, v in cmap.items():
            if part_n and (part_n in k or k in part_n):
                return v
    return None


def find_category_ids(xlsm_path: str, category_path: list[str]) -> tuple[str, str]:
    """
    在 xlsm 的 Category Tree Lookup 里找到匹配的 (MainCategory, LowestCategory)，
    返回带 ID 的完整字符串，如 ('Storage Devices (15639)', 'Portable HDD (20095)')。
    """
    if not category_path:
        return "", ""

    # 从最末级到最顶级逐层尝试匹配
    candidates = list(reversed(category_path))
    try:
        wb = openpyxl.load_workbook(xlsm_path, read_only=True, keep_vba=False)
        ws = wb["Category Tree Lookup"]
        rows = [(str(row[0]).strip(), str(row[1]).strip() if len(row) > 1 and row[1] else "")
                for row in ws.iter_rows(min_row=4, values_only=True)
                if row and row[0]]
        wb.close()
    except Exception:
        return "", ""

    for part in candidates:
        target = _norm(_strip_id(part))
        if not target:
            continue
        for main, low in rows:
            leaf = _norm(_strip_id(low.split("->")[-1] if "->" in low else low))
            if target == leaf or target in leaf or leaf in target:
                return main, low or main

    # 未匹配：返回空，调用方处理（不返回第一行避免错误类目）
    return "", ""


# ─── 字段值推断 ────────────────────────────────────────────────────────────────

def _yes_no(v: Any) -> str:
    if isinstance(v, bool):
        return "Yes" if v else "No"
    s = str(v or "").strip().lower()
    if s in {"yes", "true", "1", "y"}:
        return "Yes"
    if s in {"no", "false", "0", "n"}:
        return "No"
    return ""


def _attrs_has(draft: ListingDraft, *keywords: str) -> bool:
    blob = " ".join([
        str(draft.attributes.get(k, "")) for k in draft.attributes
    ] + [draft.title, draft.subtitle]).lower()
    return any(kw.lower() in blob for kw in keywords)


def _derive_connectivity(draft: ListingDraft) -> str:
    if _attrs_has(draft, "bluetooth"):
        return "Bluetooth"
    if _attrs_has(draft, "wireless", "2.4g", "wifi"):
        return "Wireless"
    return "Wired"


def _derive_peripheral_connectivity(draft: ListingDraft) -> str:
    """Takealot peripheral_connectivities only accepts 'Wired' or 'Wireless'.
    Bluetooth counts as Wireless."""
    conn = _derive_connectivity(draft)
    if conn in ("Bluetooth", "Wireless"):
        return "Wireless"
    return "Wired"


def _derive_usb_connectivity(draft: ListingDraft) -> str:
    """
    为 Attribute.usb_connectivity.[0] 推断一个合法值。

    合法选项（来自 loadsheet Lookup）：
        - USB 2.0
        - USB 3.0
        - USB C
        - USB Type A

    规则：
      1. 优先解析标题/属性中的明确文字（USB 3.0 / USB 2.0 / Type‑C / Type‑A）
      2. 若检测到 3.x 但未标明 2.0，则归一为 USB 3.0
      3. 若均未检测到，但类目是存储设备（Storage Devices / Portable HDD 等），默认 USB 3.0
      4. 其它类目无法判断时返回空串，让 loadsheet 忽略该列
    """
    attrs = draft.attributes or {}
    cat_path = attrs.get("_category_path") or []
    if isinstance(cat_path, str):
        cat_path = [cat_path]
    cat_blob = " ".join(str(x) for x in cat_path).lower()

    text_blob = " ".join(
        [
            draft.title or "",
            draft.subtitle or "",
            str(draft.key_features or ""),
            " ".join(str(v) for v in attrs.values()),
        ]
    ).lower()

    def has(*keywords: str) -> bool:
        return any(k.lower() in text_blob for k in keywords)

    # 明确标准文案
    if "usb 2.0" in text_blob or "usb2.0" in text_blob or "usb2 " in text_blob:
        return "USB 2.0"
    if "usb 3.0" in text_blob or "usb3.0" in text_blob or "usb3 " in text_blob or "usb 3.1" in text_blob or "usb 3.2" in text_blob:
        return "USB 3.0"

    # 接口形态
    if any(k in text_blob for k in ["type-c", "type c", "typec", "usb-c", "usb c", "c接口"]):
        return "USB C"
    if any(k in text_blob for k in ["type-a", "type a", "usb-a", "usb a"]):
        return "USB Type A"

    # 有 USB + 3.x 字样，视为 USB 3.0
    import re as _re

    if "usb" in text_blob:
        m = _re.search(r"usb\\s*([23])(?:\\.([0-9]))?", text_blob)
        if m:
            ver = m.group(1)
            if ver == "3":
                return "USB 3.0"
            if ver == "2":
                return "USB 2.0"

    # 存储设备类目，无法判断时给一个合理默认（外置硬盘通常是 USB 3.0）
    if any(k in cat_blob for k in ["storage device", "storage devices", "portable hdd", "external hard drive", "storage devices (15639)"]):
        return "USB 3.0"

    # 其它类目不强行填写，留空交给前台表单
    return ""


def _derive_hard_drive_types(draft: ListingDraft) -> str:
    """
    为 Attribute.hard_drive_types.[0] 推断合法值。

    合法选项（来自 loadsheet Lookup）包括（但不限于）：
        - Network Attached Storage (NAS)
        - Solid State Drives (SSD)
        - Surveillance
        - USB Flash Drives
        - Wireless
        - Hard Disk Drive (HDD)
        - Mechanical Hard Drive
        - eMMC
        - Dual
        - 64 GB UFS 2.2 (uMCP)
        - 128 GB UFS 2.2 (uMCP)

    规则：
      1. 根据标题/属性中的关键词区分 SSD / U 盘 / NAS / 监控盘 / 无线盘 / eMMC 等
      2. 如果属于存储设备类目（Storage Devices / Portable HDD 等），但未检测到更具体类型，
         默认填 \"Hard Disk Drive (HDD)\"，避免必填字段为空。
      3. 其它类目则返回空串，让 loadsheet 忽略该列。
    """
    attrs = draft.attributes or {}
    cat_path = attrs.get("_category_path") or []
    if isinstance(cat_path, str):
        cat_path = [cat_path]
    cat_blob = " ".join(str(x) for x in cat_path).lower()

    text_blob = " ".join(
        [
            draft.title or "",
            draft.subtitle or "",
            str(draft.key_features or ""),
            " ".join(str(v) for v in attrs.values()),
        ]
    ).lower()

    def has(*keywords: str) -> bool:
        return any(k.lower() in text_blob for k in keywords)

    # SSD / 固态盘
    if has("ssd", "固态硬盘", "固态", "pssd"):
        return "Solid State Drives (SSD)"
    # U 盘 / USB 闪存
    if has("u盘", "u 盘", "usb flash", "flash drive", "thumb drive", "pen drive"):
        return "USB Flash Drives"
    # NAS
    if has("nas", "network attached storage"):
        return "Network Attached Storage (NAS)"
    # 监控盘
    if has("监控", "surveillance", "cctv", "录像机", "nvr", "dvr"):
        return "Surveillance"
    # 无线硬盘
    if has("wireless", "wifi", "wi-fi", "无线硬盘"):
        return "Wireless"
    # eMMC / UFS
    if has("emmc", "ufs", "umcp"):
        return "eMMC"

    # 如果类目明显是存储设备（Storage Devices / Portable HDD 等），默认 HDD
    storage_like = any(
        k in cat_blob
        for k in [
            "storage devices",
            "storage device",
            "portable hdd",
            "hard drives",
            "hard drive",
            "ssd & hdd",
        ]
    ) or has("硬盘", "hard disk", "portable hdd", "external hard drive")
    if storage_like:
        return "Hard Disk Drive (HDD)"

    return ""


_HEADSET_FORM_FACTOR_MAP = {
    "in-ear": "In Ear",
    "in ear": "In Ear",
    "earbud": "In Ear",
    "earphone": "In Ear",
    "on-ear": "On Ear",
    "on ear": "On Ear",
    "over-ear": "Over Ear",
    "over ear": "Over Ear",
    "circumaural": "Over Ear",
    "supra-aural": "On Ear",
}

_VALID_HEADSET_FORM_FACTORS = {"In Ear", "On Ear", "Over Ear"}


def _derive_headset_form_factor(draft: ListingDraft) -> str:
    raw = str(draft.attributes.get("headsets_form_factor", draft.attributes.get("form_factor", ""))).strip()
    if raw in _VALID_HEADSET_FORM_FACTORS:
        return raw
    mapped = _HEADSET_FORM_FACTOR_MAP.get(raw.lower(), "")
    if mapped:
        return mapped
    # Infer from title/subtitle
    blob = (draft.title + " " + (draft.subtitle or "")).lower()
    if "over" in blob and "ear" in blob:
        return "Over Ear"
    if "on" in blob and "ear" in blob:
        return "On Ear"
    return "In Ear"  # most common default


_HEADPHONE_STYLE_MAP = {
    "in-ear": "In-Ear",
    "in ear": "In-Ear",
    "earbud": "In-Ear",
    "earphone": "In-Ear",
    "on-ear": "On Ear",
    "on ear": "On Ear",
    "over-ear": "Over Ear",
    "over ear": "Over Ear",
}

_VALID_HEADPHONE_STYLES = {"In-Ear", "On Ear", "Over Ear"}


def _derive_headphone_style(draft: ListingDraft) -> str:
    raw = str(draft.attributes.get("headphone_style", draft.attributes.get("form_factor", ""))).strip()
    if raw in _VALID_HEADPHONE_STYLES:
        return raw
    mapped = _HEADPHONE_STYLE_MAP.get(raw.lower(), "")
    if mapped:
        return mapped
    blob = (draft.title + " " + (draft.subtitle or "")).lower()
    if "over" in blob and "ear" in blob:
        return "Over Ear"
    if "on" in blob and "ear" in blob:
        return "On Ear"
    return "In-Ear"


_VALID_HEADSET_SPECIAL_FEATURES = {
    "Microphone Rotation",
    "Hands Free Calls",
    "Long Lasting Battery Life",
    "Noise Cancelling Microphone",
    "Quick Disconnect",
}


def _derive_headset_special_features(draft: ListingDraft) -> str:
    """Pick the best valid option for cellphone_headsets_special_features."""
    raw = str(draft.attributes.get("special_features", "")).strip()
    if raw in _VALID_HEADSET_SPECIAL_FEATURES:
        return raw
    blob = (draft.title + " " + (draft.subtitle or "") + " " + " ".join(
        str(v) for v in draft.attributes.values()
    )).lower()
    if "noise cancel" in blob:
        return "Noise Cancelling Microphone"
    if "long batter" in blob or "battery life" in blob:
        return "Long Lasting Battery Life"
    if "rotation" in blob and "mic" in blob:
        return "Microphone Rotation"
    if "quick disconnect" in blob:
        return "Quick Disconnect"
    return "Hands Free Calls"  # most generic valid option


def _parse_weight_g(raw: str) -> str:
    """
    从中文/混合格式重量字符串里提取克数（优先取带盒/包装重量）。
    示例：
      "单218g+带盒385g"  → "385"
      "385g"            → "385"
      "0.5kg"           → "500"
      "500"             → "500"
    返回纯数字字符串，解析失败返回原字符串。
    """
    s = str(raw or "").strip()
    if not s:
        return s
    # 优先找"带盒"/"含包装"/"含盒"后面的数字
    box_m = re.search(r"(?:带盒|含包装|含盒|box|package)[^\d]*(\d+(?:\.\d+)?)\s*(?:kg|g)?", s, re.I)
    if box_m:
        val = float(box_m.group(1))
        # 如果是 kg 单位，换算成 g
        if re.search(r"kg", s[box_m.start():box_m.end()+3], re.I):
            val = round(val * 1000)
        return str(int(val))
    # kg 单位
    kg_m = re.search(r"(\d+(?:\.\d+))\s*kg", s, re.I)
    if kg_m:
        return str(int(float(kg_m.group(1)) * 1000))
    # 所有数字里取最大值（一般最大的就是包装重量）
    nums = re.findall(r"\d+(?:\.\d+)?", s)
    if nums:
        return str(int(max(float(n) for n in nums)))
    return s


def _derive_warranty_type(draft: ListingDraft) -> str:
    wt = str(draft.attributes.get("warranty_type", "")).strip()
    if wt:
        return wt
    if _attrs_has(draft, "no warranty", "without warranty"):
        return "No Warranty"
    return "Limited"


def _derive_warranty_months(draft: ListingDraft) -> str:
    wm = str(draft.attributes.get("warranty_months", "")).strip()
    if wm and wm.isdigit():
        return wm
    return "6"


def _derive_max_power(draft: ListingDraft) -> str:
    """从属性或标题推断最大功率（瓦特），无法推断则给合理默认值。"""
    attrs = draft.attributes
    # 直接有值
    for key in ("max_power_output", "power_output", "power_w", "wattage", "watts"):
        v = str(attrs.get(key, "")).strip()
        if v and v.replace(".", "").isdigit():
            return v
    # 从标题/描述里提取数字+W
    import re as _re
    blob = draft.title + " " + (draft.subtitle or "") + " " + " ".join(str(v) for v in attrs.values())
    m = _re.search(r"(\d+)\s*[Ww](?:att)?s?\b", blob)
    if m:
        return m.group(1)
    # 根据类目给合理默认
    cat_blob = " ".join(str(v) for v in attrs.values()).lower() + draft.title.lower()
    if "subwoofer" in cat_blob:
        return "100"
    if "soundbar" in cat_blob:
        return "60"
    if "speaker" in cat_blob or "audio" in cat_blob:
        return "10"
    if "headphone" in cat_blob or "headset" in cat_blob or "earphone" in cat_blob:
        return "5"
    return "10"  # 通用默认


def _derive_country_of_origin(draft: ListingDraft) -> str:
    return str(draft.attributes.get("country_of_origin", "China")).strip() or "China"


def _parse_cm_number(raw: str) -> str:
    """
    从类似 \"9 cm\" / \"15cm\" / \"4.5 CM\" / \"90°\" 里提取纯数字部分，返回字符串。
    如果找不到数字，就返回原始字符串。
    """
    s = str(raw or "").strip()
    if not s:
        return s
    m = re.search(r"(-?\\d+(?:\\.\\d+)?)", s)
    if not m:
        return s
    return m.group(1)


def _derive_product_dimensions(draft: ListingDraft) -> tuple[str, str, str]:
    """
    推断产品实体尺寸（宽/长/高，单位 cm）。
    优先级：
      1. 1688 直接提供的产品尺寸字段
      2. 1688 包装尺寸（略微缩小作估算）
      3. 用 LLM 根据产品标题+属性估算
      4. 硬编码类目默认值
    返回 (width, length, height) 字符串三元组。
    """
    attrs = draft.attributes

    # 1. 直接尺寸
    w = str(attrs.get("width_cm", "")).strip()
    l = str(attrs.get("length_cm", "")).strip()
    h = str(attrs.get("height_cm", "")).strip()
    if w and l and h:
        return w, l, h

    # 2. 包装尺寸缩小 20% 作估算
    pw = str(attrs.get("packaged_width", "")).strip()
    pl = str(attrs.get("packaged_length", "")).strip()
    ph = str(attrs.get("packaged_height", "")).strip()
    if pw and pl and ph:
        try:
            return (
                str(round(float(pw) * 0.8, 1)),
                str(round(float(pl) * 0.8, 1)),
                str(round(float(ph) * 0.8, 1)),
            )
        except ValueError:
            pass

    # 3. LLM 估算
    try:
        from .llm import _call_llm_json, is_llm_available
        if is_llm_available():
            attrs_str = ", ".join(f"{k}: {v}" for k, v in list(attrs.items())[:15] if v)
            prompt = (
                f"Product: {draft.title}\nAttributes: {attrs_str}\n"
                "Estimate the product's physical assembled dimensions in centimeters. "
                "Return JSON: {\"width\": <number>, \"length\": <number>, \"height\": <number>}. "
                "Numbers only, no units. Be realistic for this type of product."
            )
            result = _call_llm_json(prompt, temperature=0.2)
            lw = str(result.get("width", "")).strip()
            ll = str(result.get("length", "")).strip()
            lh = str(result.get("height", "")).strip()
            if lw and ll and lh:
                return lw, ll, lh
    except Exception:
        pass

    # 4. 类目默认值
    blob = (draft.title + " " + " ".join(str(v) for v in attrs.values())).lower()
    if "subwoofer" in blob:
        return "25", "30", "30"
    if "soundbar" in blob:
        return "60", "10", "8"
    if "speaker" in blob or "audio" in blob:
        return "15", "12", "10"
    if "headphone" in blob or "headset" in blob:
        return "18", "20", "8"
    if "earphone" in blob or "earbud" in blob:
        return "6", "4", "3"
    return "15", "10", "8"  # 通用默认


def _split_colours(raw: str) -> list[str]:
    """
    把 1688/LLM 里类似 \"Red / Blue / Black / White\" 这样的颜色串拆成单个颜色 token。
    仅做简单分割，不尝试翻译；由调用方决定取第几个作为主/副颜色。
    """
    s = str(raw or "").strip()
    if not s:
        return []
    import re as _re
    parts = _re.split(r"[/,、|]+", s)
    out: list[str] = []
    for p in parts:
        t = p.strip()
        if not t:
            continue
        # 去掉多余空格，比如 \"Red  \" -> \"Red\"
        out.append(t)
    return out


def _derive_sku(draft: ListingDraft, source: ProductSource) -> str:
    mn = str(draft.attributes.get("model_number", "")).strip()
    if mn:
        return re.sub(r"[^\w\-]", "", mn)[:40]
    # 从标题取前几个单词
    words = re.findall(r"[A-Z0-9]+", draft.title.upper())
    return ("-".join(words[:3]) or "SKU-AUTO")[:40]


def _apply_portal_field_overrides(
    row_values: dict,
    attrs: dict,
    xlsm_path: str,
) -> dict:
    """
    通用 portal 字段透传：把 draft.attributes 里以 portal 字段标签为 key 的值
    自动映射到对应 xlsm 列，无需为每个类目单独硬编码。

    匹配规则：将 portal label 归一化（小写+下划线），在 xlsm 列 key 中查找包含该字符串的列。
    例如：
      "Screen Size"  -> norm "screen_size" -> matches "Attribute.screen_size.value"
      "Primary Sports Type" -> norm "primary_sports_type" -> matches "Attribute.primary_sports_type"
      "Operating System Compatibility" -> norm "operating_system_compatibility"
                                       -> matches "Attribute.operating_system_compatibility.[0]"
    已有值的列不会被覆盖（只填空白列）。
    """
    import re as _re
    try:
        wb = openpyxl.load_workbook(xlsm_path, read_only=True, keep_vba=False)
        sheet_name = "Loadsheet" if "Loadsheet" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        header_keys = [str(c.value).strip() if c.value else "" for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
        wb.close()
    except Exception:
        return row_values

    # 建立归一化 key → xlsm 列 key 的反查表
    def _norm_col(col_key: str) -> str:
        """Attribute.primary_sports_type.[0] -> primary_sports_type"""
        s = col_key.removeprefix("Attribute.")
        s = _re.sub(r"\.\[?\d+\]?$", "", s)   # remove .[0]
        s = _re.sub(r"\.value$", "", s)         # remove .value
        return s.replace(".", "_").lower()

    norm_to_col: dict[str, str] = {}
    for col_key in header_keys:
        if col_key:
            norm_to_col[_norm_col(col_key)] = col_key

    # 特例映射：portal 标签与 loadsheet 机器列名不一致的字段（手工指定映射）。
    # key = portal label 归一化后的字符串，value = xlsm 列归一化后的字符串
    # 根因：Takealot portal 表单标签和 xlsm 列名常有以下差异：
    #   - 单复数：Connectivity → connectivities, Game Platform → game_platforms
    #   - is_ 前缀：Ergonomic → is_ergonomic, Rechargeable → is_rechargeable
    #   - 颜色别名：Colour → color.main（color_main）
    #   - 全名缩短：Main Material/Fabric → materials
    #   - 结构差异：Warranty → warranty.type（warranty_type）
    _LABEL_ALIAS = {
        # ── 旧有别名 ──
        "ingredients":                  "active_ingredients",
        "healthcare_product_format":    "format_healthcare",
        "has_expiry_date":              "has_expiry_days",
        # ── 单复数 ──
        "game_platform":                "game_platforms",           # "Game Platform" → Attribute.game_platforms.[0]
        "connectivity":                 "connectivities",           # "Connectivity" → Attribute.connectivities.[0]
        "peripheral_connectivity":      "peripheral_connectivities",# "Peripheral Connectivity" → Attribute.peripheral_connectivities.[0]
        "compatible_device":            "compatible_devices",       # "Compatible Device(s)" → Attribute.compatible_devices
        # ── is_ 前缀缺失 ──
        "ergonomic":                    "is_ergonomic",             # "Ergonomic" → Attribute.is_ergonomic
        "rechargeable":                 "is_rechargeable",          # "Rechargeable" → Attribute.is_rechargeable
        "water_resistant":              "is_water_resistant",       # "Water Resistant" → Attribute.is_water_resistant
        "waterproof":                   "is_waterproof",            # "Waterproof" → Attribute.is_waterproof
        "portable":                     "is_portable",              # "Portable" → Attribute.is_portable
        "wireless":                     "is_wireless",              # "Wireless" → Attribute.is_wireless
        # ── 颜色别名 ──
        "colour":                       "color_main",               # "Colour" (probe) → color.main
        # ── 材质 ──
        "main_material_fabric":         "materials",                # "Main Material/Fabric" → Attribute.materials.[0]
        "main_strap_material":          "materials",                # "Main Strap Material" → Attribute.materials.[0]
        # ── 结构差异 ──
        "warranty":                     "warranty_type",            # "Warranty" → Attribute.warranty.type
        # ── 文本字段别名（portal label → xlsm plain column） ──
        "product_title":                "title",                    # "Product Title" → title
        "key_selling_features":         "description",              # "Key Selling Features" → description
        # ── 包装尺寸 / 重量（portal label 带单位，xlsm 列是完整路径） ──
        "packaged_height_cm":           "merchant_packaged_dimensions_height",  # "Packaged Height (cm)"
        "packaged_length_cm":           "merchant_packaged_dimensions_length",  # "Packaged Length (cm)"
        "packaged_width_cm":            "merchant_packaged_dimensions_width",   # "Packaged Width (cm)"
        "packaged_weight_g":            "merchant_packaged_weight",             # "Packaged Weight (g)"
        # ── 数值字段 ──
        "device_storage_capacity_value":"device_storage_capacity",  # "Device Storage Capacity value"
        "office_user_type":             "by_office_user",           # "Office User Type" → Attribute.by_office_user.[0]
    }

    # ── 通用"选项集匹配"：当标签别名找不到列时，用 probe 的选项集与 xlsm LOOKUP 对比 ──
    # 原理：portal 表单里每个下拉字段的有效值集合与 xlsm LOOKUP 的值集合完全一致，
    # 通过 frozenset 精确匹配，完全不依赖标签名，对新类目自动生效。
    # 限制：只在选项数 ≥ 3 时启用，避免 Yes/No 这类通用集合产生歧义。
    _options_to_col: dict[frozenset, str] = {}
    try:
        _xlsm_lookups = _extract_all_valid_values(xlsm_path)
        for _lk_name, _lk_vals in _xlsm_lookups.items():
            if not _lk_vals or len(_lk_vals) < 3:
                continue
            # LOOKUP_Attribute_by_office_user_0 → by_office_user
            _lk_norm = _lk_name.lower()
            _lk_norm = _re.sub(r"^lookup_attribute_", "", _lk_norm)
            _lk_norm = _re.sub(r"^lookup_variant_", "", _lk_norm)
            _lk_norm = _re.sub(r"_\d+$", "", _lk_norm)
            if _lk_norm in norm_to_col:
                _opts_set = frozenset(str(v).strip().lower() for v in _lk_vals if v)
                if _opts_set and _opts_set not in _options_to_col:  # 第一个匹配到的列优先
                    _options_to_col[_opts_set] = norm_to_col[_lk_norm]
    except Exception:
        pass

    # 从 _probe_fields 建立 label → options 的查找表，用于写入前验证下拉值
    probe_options: dict[str, list[str]] = {}
    pf_raw = attrs.get("_probe_fields")
    if isinstance(pf_raw, list):
        for pf in pf_raw:
            lbl = str(pf.get("label", "")).strip()
            opts = pf.get("options") or []
            if lbl and opts:
                probe_options[lbl.lower()] = [str(o).strip() for o in opts if str(o).strip()]

    result = dict(row_values)
    for attr_key, attr_val in attrs.items():
        if not attr_val or not isinstance(attr_key, str):
            continue
        if attr_key.startswith("_"):
            continue
        # 只处理"看起来像 portal label"的 key（包含空格 或 首字母大写）
        if not (" " in attr_key or (attr_key and attr_key[0].isupper())):
            continue
        # 归一化：把 / 替换为空格（不能直接删掉，否则 "Material/Fabric" → "materialfabric"）
        # 再去掉 & ' " ( ) 等符号，空格/连字符 → 下划线
        norm = re.sub(r"[/]+", " ", attr_key.lower())
        norm = re.sub(r"[&'\"()]+", "", norm).strip()
        norm = re.sub(r"[\s\-]+", "_", norm)
        # 优先使用别名映射
        norm_lookup = _LABEL_ALIAS.get(norm, norm)
        col_key = norm_to_col.get(norm_lookup)

        # 标签匹配失败时，尝试通过"选项集"自动找对应列（适用于新类目无需手动加别名）
        if not col_key:
            field_opts = probe_options.get(attr_key.lower())
            if field_opts and len(field_opts) >= 3:
                opts_set = frozenset(o.strip().lower() for o in field_opts)
                col_key = _options_to_col.get(opts_set)
                if col_key:
                    print(f"[csv_exporter] 📍 选项集自动匹配：{attr_key!r} → {col_key}")

        if not col_key:
            # 只对 probe 必填字段打印映射缺失警告，可选字段和已被其他地方处理的字段静默跳过
            if isinstance(pf_raw, list):
                req_norms = {
                    re.sub(r"[\s\-/]+", "_", re.sub(r"[&'\"()]+", "", str(pf.get("label","")).strip().lower()))
                    for pf in pf_raw if pf.get("required")
                }
                if norm in req_norms:
                    print(f"[csv_exporter] ⚠️  必填 portal 字段 {attr_key!r} 未找到对应 xlsm 列（norm={norm!r}），请在 _LABEL_ALIAS 中添加映射")
            continue
        # portal 字段值（用户在预览里显式选择的）优先于 build_row_values 自动推断的值
        # 不再有 "已有值不覆盖" 的保护，portal 值总是覆盖旧值
        val_str = str(attr_val).strip()
        # 如果该字段有 dropdown options，验证值必须在列表里，否则跳过（不写无效值）
        opts = probe_options.get(attr_key.lower())
        if opts:
            opts_lower = [o.lower() for o in opts]
            if val_str.lower() not in opts_lower:
                print(f"[csv_exporter] 跳过无效下拉值 {attr_key!r}={val_str!r}（不在选项列表中）")
                continue
            # 用原始大小写
            idx = opts_lower.index(val_str.lower())
            val_str = opts[idx]
        result[col_key] = val_str

    return result


def _derive_wearable_features(draft: ListingDraft) -> dict:
    """从标题 / 描述中检测智能可穿戴功能，自动填写 Activity Tracker / Smart Watch 必填列。"""
    blob = " ".join([
        str(draft.title or ""),
        str(draft.subtitle or ""),
        str(draft.key_features or ""),
        str(draft.source_url or ""),
    ]).lower()
    attrs = draft.attributes or {}

    def _feat(*keywords: str) -> str:
        user_val = ""
        for k in keywords:
            # Check both underscore format ("music_player") and portal label format ("Music Player")
            v = str(attrs.get(k.replace(" ", "_"), "") or attrs.get(k, "") or attrs.get(k.title(), "")).strip()
            if v:
                user_val = v
                break
        if user_val:
            return user_val
        return "Yes" if any(kw in blob for kw in keywords) else "No"

    return {
        "Attribute.has_heart_rate_monitor":         _feat("heart rate", "heart-rate", "heartrate"),
        "Attribute.has_calorie_counter":             _feat("calorie", "calories"),
        "Attribute.has_step_counter":                _feat("step counter", "pedometer", "step count"),
        "Attribute.has_sleep_monitor":               _feat("sleep monitor", "sleep tracking", "sleep quality"),
        "Attribute.has_distance_tracking":           _feat("distance tracking", "distance tracker"),
        "Attribute.has_gps":                         _feat("gps", "global positioning"),
        "Attribute.has_glonass":                     _feat("glonass"),
        "Attribute.has_altimeter":                   _feat("altimeter", "altitude"),
        "Attribute.has_music_player":                _feat("music player", "plays music", "music control"),
        "Attribute.has_smart_notification_technology": _feat("notification", "smart notification"),
        "Attribute.has_2way_calling":                _feat("phone call", "two-way", "2-way calling"),
        "Attribute.has_sos_panic_button":            _feat("sos", "panic button", "emergency alert"),
        "Attribute.has_topo_mapping_technology":     _feat("topo", "topographic map"),
        "Attribute.has_safe_zone_technology":        _feat("safe zone", "geofence"),
        "Attribute.has_monitoring_capabilities":     _feat("monitoring", "health monitor", "vital"),
        "Attribute.has_sas_messaging_technology":    _feat("sas messaging", "satellite"),
        "Attribute.touchscreen_enabled":             _feat("touchscreen", "touch screen"),
        "Attribute.has_bluetooth":                   _feat("bluetooth"),
        "Attribute.is_water_resistant":              _feat("water resistant", "water-resistant", "ip6", "ipx"),
        "Attribute.is_waterproof":                   _feat("waterproof", "ip68", "ip67"),
        "Attribute.is_portable":                     "Yes",
        "Attribute.customer_registration_required":  _feat("registration required", "account required"),
    }


# ─── 核心：构建列 key → 值 的映射 ─────────────────────────────────────────────

def build_row_values(
    draft: ListingDraft,
    source: ProductSource,
    main_category: str,
    lowest_category: str,
    image_urls: list[str],
) -> dict[str, Any]:
    """
    返回 {column_key: value} 字典，覆盖所有能自动推断的字段。
    未知字段留给调用方处理或留空。
    """
    attrs = draft.attributes
    sku = _derive_sku(draft, source)

    row: dict[str, Any] = {
        # ── 基础结构 ──
        "Variant.ProductVariant": "Product without Variants",
        "SKU": sku,
        "Variant.ProductCode": "",          # 无变体时为空

        # ── 类目 ──
        "TopCategory": main_category,
        "Category":    lowest_category,

        # ── 识别 ──
        "ProductID.Value": str(attrs.get("barcode", "") or attrs.get("ean", "") or ""),

        # ── 内容 ──
        "title":                       draft.title[:100],
        "subtitle":                    draft.subtitle[:100] if draft.subtitle else "",
        "description":                 draft.key_features[:2000] if draft.key_features else "",
        "Attribute.whats_in_the_box":  ", ".join(draft.whats_in_box) if draft.whats_in_box else "",
        # Brand 默认留空，避免随意写品牌；如需品牌由人工在 loadsheet 中补充
        "Brand":                       "",

        # ── 型号 ──
        "Attribute.model_number": str(attrs.get("model_number", "") or sku),
        "Attribute.model_name":   str(attrs.get("model_name",   "") or ""),

        # ── 颜色 / 材质 ──
        # 注：color.main 由 1688 预填，但若预览中用户通过 "Colour" 字段选了合规值，
        #     _apply_portal_field_overrides 会覆盖此处的值。
        "color.main":               _non_placeholder(attrs.get("colour", "") or attrs.get("color", "")),
        "color.secondary":          _non_placeholder(attrs.get("secondary_colour", "")),
        "color.name":               _non_placeholder(attrs.get("colour_name", "") or attrs.get("colour", "")),
        # 注：不在此处写 Attribute.materials.[0]，原因：
        #   1688/LLM 粗略材质（如 "ABS Plastic"）常不在 Takealot 的下拉合规列表里会导致驳回。
        #   若 probe 探测到 "Main Material/Fabric" 字段，用户在预览中选择后，
        #   _apply_portal_field_overrides 会通过 "main_material_fabric"→"materials" 别名写入。

        # ── 连接 ──
        "Attribute.connectivities.[0]":            _derive_connectivity(draft),
        "Attribute.peripheral_connectivities.[0]": _derive_peripheral_connectivity(draft),
        "Attribute.usb_connectivity.[0]":          _derive_usb_connectivity(draft),
        "Attribute.bluetooth_version":             str(attrs.get("bluetooth_version", "")),
        "Attribute.is_wireless": _yes_no(_attrs_has(draft, "wireless", "bluetooth", "2.4g")),
        "Attribute.has_bluetooth": _yes_no(_attrs_has(draft, "bluetooth")),

        # ── 耳机专用 ──
        # Valid headsets_form_factor options: 'In Ear', 'On Ear', 'Over Ear'
        "Attribute.headsets_form_factor": _derive_headset_form_factor(draft),
        # Valid headphone_styles options: 'In-Ear', 'On Ear', 'Over Ear'
        "Attribute.headphone_styles.[0]": _derive_headphone_style(draft),
        # Valid special features: Microphone Rotation, Hands Free Calls,
        #   Long Lasting Battery Life, Noise Cancelling Microphone, Quick Disconnect
        "Attribute.cellphone_headsets_special_features": _derive_headset_special_features(draft),
        "Attribute.has_noise_cancelling": _yes_no(attrs.get("noise_cancelling", False)),
        # Valid: Yes, No — default No (most headsets don't have remote)
        "Attribute.has_integrated_remote": _yes_no(attrs.get("has_integrated_remote", False)),

        # ── 音箱专用 ──
        # Valid: Bookshelf, Ceiling, Floorstanding, Wall Mount, Table Mounted, Clip On
        "Attribute.speaker_mountings.[0]": str(attrs.get("speaker_mounting", "Table Mounted")),
        # Valid: 2.5 mm mono (TS), 3.5 mm stereo (TRS), 6.35 mm (1/4 inch) stereo (TRS), No Jack
        "Attribute.audio_jacks.[0]": str(attrs.get("audio_jack", "No Jack")),
        # Valid: Computers, DVD Players, Gaming Consoles, Music Players, Smartphones,
        #        Surround Sound Systems, Televisions, Tablets, Universal, Wearable Speaker
        "Attribute.speaker_recommended_uses.[0]": str(attrs.get("speaker_recommended_use", "Universal")),
        # Valid: Hi - Fi Sound, Dynamic RGB Lighting, Multiple Ways to Connect, Quiet Standby, Energy Saving
        "Attribute.speaker_features": str(attrs.get("speaker_features", "Multiple Ways to Connect")),
        # Max Power Output (watts) — 从属性或描述推断，无则估算
        "Attribute.max_power_output.value": _derive_max_power(draft),

        # ── 产品实体尺寸（Assembled Dimensions，Takealot 必填）──
        # 优先用 1688 的产品尺寸；无则用包装尺寸估算；仍空则 LLM 推断；最终给类目默认值
    }
    _dim_w, _dim_l, _dim_h = _derive_product_dimensions(draft)
    row.update({
        "Attribute.product_dimensions.width":  _dim_w,
        "Attribute.product_dimensions.length": _dim_l,
        "Attribute.product_dimensions.height": _dim_h,
    })

    # ── 颜色字段清洗：只保留单一主色 + 可选副色 ──
    # 1688/LLM 常给出 \"Red / Blue / Black / White\" 这样的串，这里拆分后：
    #   - 主色：第 1 个 token（例如 \"Red\"）
    #   - 副色：第 2 个 token（如果存在且 secondary 目前为空）
    main_colour_raw = str(row.get("color.main", "")).strip()
    if main_colour_raw:
        tokens = _split_colours(main_colour_raw)
        if tokens:
            row["color.main"] = tokens[0]
            # Colour Name 缺省时沿用主色
            if not str(row.get("color.name", "")).strip():
                row["color.name"] = tokens[0]
            # Secondary Colour 为空且有第二个 token 时，用作副色
            if not str(row.get("color.secondary", "")).strip() and len(tokens) >= 2:
                row["color.secondary"] = tokens[1]

    # Brand 里如果意外塞进了类似 \"Red / Blue / Black\" 这种值，视为错误数据，直接清空，
    # 避免把颜色串当品牌名。
    brand_raw = str(row.get("Brand", "")).strip()
    if brand_raw and ("/" in brand_raw or "," in brand_raw):
        row["Brand"] = ""

    row.update({

        # ── 是/否 通用属性 ──
        "Attribute.is_rechargeable":   _yes_no(attrs.get("rechargeable", False)),
        "Attribute.fast_charging":     _yes_no(attrs.get("fast_charging", False)),
        "Attribute.is_portable":       _yes_no(attrs.get("portable", True)),
        "Attribute.is_foldable":       _yes_no(attrs.get("foldable", False)),
        "Attribute.is_water_resistant":_yes_no(attrs.get("water_resistant", False)),
        "Attribute.is_waterproof":     _yes_no(attrs.get("waterproof", False)),
        "Attribute.is_adjustable":     _yes_no(attrs.get("adjustable", False)),
        "Attribute.is_lightweight":    _yes_no(attrs.get("lightweight", True)),
        "Attribute.proudly_south_african":           "No",
        "Attribute.customer_registration_required":  "No",

        # ── 电池 ──
        "Attribute.battery_type":           str(attrs.get("battery_type", "")),
        "Attribute.battery_capacity.value": str(attrs.get("battery_capacity", "")),
        "Attribute.battery_size.[0]":       str(attrs.get("battery_size", "")),

        # ── 原产地 ──
        "Attribute.country_of_origin.[0]": _derive_country_of_origin(draft),

        # ── 保修 ──
        "Attribute.warranty.type":         _derive_warranty_type(draft),
        "Attribute.warranty.period.value": _derive_warranty_months(draft),

        # ── 包装尺寸重量（用 or 跳过空字符串，无数据则留空不填默认值）──
        "Attribute.merchant_packaged_dimensions.width":  str(attrs.get("packaged_width")  or attrs.get("width_cm")  or ""),
        "Attribute.merchant_packaged_dimensions.length": str(attrs.get("packaged_length") or attrs.get("length_cm") or ""),
        "Attribute.merchant_packaged_dimensions.height": str(attrs.get("packaged_height") or attrs.get("height_cm") or ""),
        "Attribute.merchant_packaged_weight.value":      _parse_weight_g(str(attrs.get("packaged_weight") or attrs.get("weight_g") or "")),

        # ── 智能可穿戴设备专用功能（Activity Trackers / Smart Watches）──
        # 从标题、副标题、描述中检测功能关键词
        **_derive_wearable_features(draft),
    })

    # ── 图片 URL（最多20张）──
    for i, url in enumerate(image_urls[:20], start=1):
        key = "Images.image_url_1" if i == 1 else f"Images.image_url_{i}"
        row[key] = url

    return row


# ─── 读取 loadsheet 所有下拉选项 ─────────────────────────────────────────────

def _extract_all_valid_values(xlsm_path: str) -> dict[str, list[str]]:
    """
    从 xlsm 的 Lookup sheet 和 named ranges 提取所有下拉列的合法值。
    返回 {column_key: [valid, values, ...]}
    """
    import re as _re
    valid: dict[str, list[str]] = {}
    try:
        wb = openpyxl.load_workbook(xlsm_path, keep_vba=False)
        if "Lookup" not in wb.sheetnames:
            wb.close()
            return valid
        lws = wb["Lookup"]
        # 遍历所有 named ranges，找 LOOKUP_Attribute_* 格式
        for name, defn in wb.defined_names.items():
            if not name.startswith("LOOKUP_Attribute_") and not name.startswith("LOOKUP_Variant_"):
                continue
            m = _re.search(r"Lookup!\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)", defn.attr_text)
            if not m:
                continue
            rng = f"{m.group(1)}{m.group(2)}:{m.group(3)}{m.group(4)}"
            try:
                vals = [c.value for row in lws[rng] for c in row if c.value is not None]
                if vals:
                    # name like LOOKUP_Attribute_speaker_features → key guess
                    key_guess = name.replace("LOOKUP_Attribute_", "Attribute.").replace("LOOKUP_Variant_", "Variant.")
                    key_guess = key_guess.replace("_", ".", 1)  # first underscore → dot
                    valid[name] = vals
            except Exception:
                pass
        wb.close()
    except Exception:
        pass
    return valid


def _get_dropdown_values_for_key(col_key: str, all_valid: dict[str, list[str]]) -> list[str] | None:
    """给定列 key，从 all_valid 里查找对应的合法值列表。"""
    # col_key 示例: "Attribute.speaker_mountings.[0]" 或 "Main Material/Fabric"
    # named range 示例: "LOOKUP_Attribute_speaker_mountings_0"
    # 先用原始规则规范化（处理 Attribute.xxx 格式）
    normalized = col_key.replace("Attribute.", "").replace("Variant.", "").replace(".[", "_").replace("]", "").replace(".", "_")
    # 再把剩余的空格/斜线/特殊字符也转为下划线（处理 "Main Material/Fabric" 这类列名）
    normalized_loose = re.sub(r"[^a-z0-9]+", "_", normalized.lower()).strip("_")
    target_attr = f"LOOKUP_Attribute_{normalized}"
    target_variant = f"LOOKUP_Variant_{normalized}"
    for name, vals in all_valid.items():
        name_lower = name.lower()
        if name_lower == target_attr.lower() or name_lower == target_variant.lower():
            return vals
        # 宽松匹配：用下划线规范化后比较（处理空格/斜线等分隔符差异）
        name_loose = re.sub(r"[^a-z0-9]+", "_", name_lower).strip("_")
        if normalized_loose and normalized_loose in name_loose:
            return vals
    return None


# ─── AI 补填空白必填字段 ──────────────────────────────────────────────────────

def ai_fill_missing_fields(
    row_values: dict[str, Any],
    xlsm_path: str,
    draft: "ListingDraft",
    source: "ProductSource",
) -> dict[str, Any]:
    """
    读出 xlsm 所有下拉列的合法值，找出 row_values 里为空的字段，
    批量让 LLM 根据产品信息选择最合适的值填入。
    """
    try:
        from .llm import _call_llm_json, is_llm_available
    except Exception:
        return row_values

    if not is_llm_available():
        return row_values

    all_valid = _extract_all_valid_values(xlsm_path)
    if not all_valid:
        return row_values

    # 占位符值视为未填（不会通过下拉校验）
    def _is_empty_or_placeholder(v: str) -> bool:
        return not v.strip() or v.strip().lower() in _PLACEHOLDER_STRS

    # 从 xlsm 读所有列 key（含未被 build_row_values 写入的列，如 case_cover_type）
    try:
        _wb_hdr = openpyxl.load_workbook(xlsm_path, read_only=True, keep_vba=False)
        _ws_hdr = _wb_hdr["Loadsheet"] if "Loadsheet" in _wb_hdr.sheetnames else _wb_hdr[_wb_hdr.sheetnames[0]]
        all_col_keys = [str(c.value).strip() if c.value else "" for c in list(_ws_hdr.iter_rows(min_row=1, max_row=1))[0]]
        _wb_hdr.close()
    except Exception:
        all_col_keys = list(row_values.keys())

    # ── 用 portal 探测字段过滤，只处理探测到的字段 ──────────────────────────
    # 没有 probe 数据时直接跳过，避免把所有类目的属性字段都乱填
    probe_fields: list[dict] = []
    if draft.attributes:
        pf_raw = draft.attributes.get("_probe_fields")  # type: ignore[call-overload]
        if isinstance(pf_raw, list):
            probe_fields = pf_raw

    if not probe_fields:
        print("[csv_exporter] 无 portal 探测数据，跳过 AI 补填（避免乱填无关字段）")
        return row_values

    if probe_fields:
        # 只针对「必填」字段做 AI 补填：构建必填字段标签的规范化集合
        probe_required_norms: set[str] = set()
        for pf in probe_fields:
            if not pf.get("required"):
                continue
            lbl = str(pf.get("label", "") or "")
            norm = re.sub(r"[^a-z0-9]+", "_", lbl.lower()).strip("_")
            if norm:
                probe_required_norms.add(norm)

        # 如果探测结果里没有任何必填字段，就直接跳过 AI 补填，避免乱填
        if not probe_required_norms:
            print("[csv_exporter] portal 探测无必填字段，跳过 AI 补填")
            return row_values

        def _col_key_in_probe(col_key: str) -> bool:
            # 提取 Attribute.xxx.value 的中间部分，其余直接规范化整个 key
            parts = col_key.split(".")
            if len(parts) >= 3 and parts[0].lower() == "attribute":
                inner = ".".join(parts[1:-1])
            else:
                inner = col_key
            inner_norm = re.sub(r"[^a-z0-9]+", "_", inner.lower()).strip("_")
            for pnorm in probe_required_norms:
                if pnorm == inner_norm or pnorm in inner_norm or inner_norm in pnorm:
                    return True
            return False

        all_col_keys = [k for k in all_col_keys if _col_key_in_probe(k)]
        print(f"[csv_exporter] probe 过滤后列数：{len(all_col_keys)}（共探测 {len(probe_fields)} 个字段）")

    # 找出空白、占位符、或值不在下拉选项里的字段（扫全部列，不只是已有 key）
    missing: dict[str, list[str]] = {}
    for col_key in all_col_keys:
        if not col_key:
            continue
        opts = _get_dropdown_values_for_key(col_key, all_valid)
        if not opts:
            continue
        current_val = str(row_values.get(col_key, "")).strip()
        # 需要 AI 处理：空/占位符，或当前值不在合法选项里（大小写不敏感比较）
        opts_lower = {o.lower() for o in opts}
        if _is_empty_or_placeholder(current_val) or current_val.lower() not in opts_lower:
            missing[col_key] = opts

    if not missing:
        return row_values

    # 构建产品摘要
    product_summary = (
        f"Product: {draft.title}\n"
        f"Subtitle: {draft.subtitle}\n"
        f"Category: {row_values.get('Category', '')}\n"
        f"Description: {(draft.key_features or '')[:500]}\n"
        f"Attributes: {dict(list(draft.attributes.items())[:20])}\n"
    )

    # 每次最多批量处理 20 个字段（避免 prompt 过长）
    result = dict(row_values)
    missing_items = list(missing.items())

    for i in range(0, len(missing_items), 20):
        batch = missing_items[i:i+20]
        fields_json = {k: v for k, v in batch}

        prompt = (
            "You are a product data specialist for Takealot South Africa.\n"
            "Given the product information below, choose the BEST value for each field "
            "from the provided options list. Return ONLY a JSON object mapping field keys to chosen values.\n\n"
            f"PRODUCT:\n{product_summary}\n\n"
            f"FIELDS TO FILL (key: [valid options]):\n{json.dumps(fields_json, indent=2)}\n\n"
            "Rules:\n"
            "- Choose exactly one value from each field's options list\n"
            "- If unsure, pick the most generic/universal option\n"
            "- Return only the JSON object, no explanation\n"
        )

        try:
            filled = _call_llm_json(prompt, temperature=0.1)
            for k, v in filled.items():
                if k not in missing:
                    continue  # 只更新 missing 里的字段（空/无效），不覆盖已有合法值
                # Validate that v is in the options list
                opts = _get_dropdown_values_for_key(k, all_valid)
                if opts and v in opts:
                    result[k] = v
                elif opts:
                    result[k] = opts[0]  # fallback to first valid option
        except Exception:
            # If LLM fails for a batch, use first valid option as fallback
            for k, opts in batch:
                if k in missing and opts:
                    result[k] = opts[0]

    filled_count = sum(1 for k, _ in missing_items if str(result.get(k, "")).strip())
    if filled_count:
        print(f"[csv_exporter] AI 补填了 {filled_count}/{len(missing_items)} 个空白字段")

    return result


# ─── 写入 xlsm ────────────────────────────────────────────────────────────────

def _read_header_keys(ws) -> list[str]:
    """读取第1行的列 key 列表（机器名）。"""
    row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    return [str(v).strip() if v else "" for v in row1]


def write_to_xlsm(
    xlsm_template: str,
    row_values: dict[str, Any],
    output_path: Path,
    data_start_row: int = 5,
    required_portal_norms: set[str] | None = None,
) -> None:
    """
    复制 xlsm 模板，在第 data_start_row 行写入数据，保存到 output_path。
    保留原模板的格式和数据验证，只填数据行。
    required_portal_norms: probe 必填字段归一化标签集合（如 {'office_user_type', 'by_office_user', ...}）；
    对应 xlsm 列为空时标橙底色，方便用户识别还有哪些必填字段没填。
    """
    import re as _re
    from openpyxl.styles import PatternFill

    shutil.copy2(xlsm_template, str(output_path))

    # openpyxl 写模式（keep_vba=True 保留宏）
    wb = openpyxl.load_workbook(str(output_path), keep_vba=True)
    sheet_name = "Loadsheet" if "Loadsheet" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    # 读列索引
    header_keys = _read_header_keys(ws)

    # 归一化 xlsm 列 key（与 _apply_portal_field_overrides 中的 _norm_col 保持一致）
    def _nc(col_key: str) -> str:
        s = col_key.removeprefix("Attribute.")
        s = _re.sub(r"\.\[?\d+\]?$", "", s)
        s = _re.sub(r"\.value$", "", s)
        return s.replace(".", "_").lower()

    # 橙底色：必填且为空的单元格
    _fill_req = PatternFill(start_color="FFCC80", end_color="FFCC80", fill_type="solid")

    # 先用空字符串覆盖整个数据行（防止有残留数据）
    for col_idx in range(1, len(header_keys) + 1):
        ws.cell(row=data_start_row, column=col_idx, value="")

    # 写数据行，并对必填空字段标橙底色
    for col_idx, key in enumerate(header_keys, start=1):
        if not key:
            continue
        value = row_values.get(key, "")
        cell = ws.cell(row=data_start_row, column=col_idx, value=str(value) if value != "" else "")
        # 如果有必填标注且值为空，检查该列是否对应某个必填 portal 字段
        if required_portal_norms and not str(value).strip():
            col_norm = _nc(key)
            if col_norm in required_portal_norms:
                cell.fill = _fill_req

    wb.save(str(output_path))
    wb.close()


# ─── 主入口 ───────────────────────────────────────────────────────────────────

def generate_loadsheet(
    draft: ListingDraft,
    source: ProductSource,
    run_dir: Path,
    image_urls: list[str] | None = None,
) -> Path | None:
    """
    生成 Takealot loadsheet xlsm 文件，保存到 run_dir。

    参数：
        draft:      ListingDraft（LLM 生成的产品草稿）
        source:     ProductSource（1688 抓取的原始数据）
        run_dir:    本次运行目录（output/runs/xxx/）
        image_urls: 图片 URL 列表（优先用；若 None 则从 source 取）

    返回：
        生成的 xlsm 路径，失败返回 None。
    """
    category_path = draft.attributes.get("_category_path") or source.category_path or []

    # 1. 找对应 loadsheet
    xlsm_path = find_loadsheet(category_path)

    # 如果中文路径匹配失败，先尝试翻译为英文再匹配
    if not xlsm_path:
        en_path = _translate_zh_category(category_path)
        if en_path != category_path:
            xlsm_path = find_loadsheet(en_path)
            if xlsm_path:
                category_path = en_path

    # 再尝试读同目录下 portal_result.json 里已解析的英文路径
    if not xlsm_path:
        portal_result = run_dir / "portal_result.json"
        if portal_result.exists():
            try:
                import json as _json
                pr = _json.loads(portal_result.read_text(encoding="utf-8"))
                en_path = pr.get("category_path_used") or []
                if en_path:
                    xlsm_path = find_loadsheet(en_path)
                    if xlsm_path:
                        category_path = en_path
            except Exception:
                pass

    if not xlsm_path:
        print(f"[csv_exporter] ✗ 找不到对应 loadsheet，类目路径：{category_path}")
        # ── LLM 自动映射：让 AI 从现有类目列表中选最匹配的 ──
        xlsm_path = _llm_map_category(category_path, draft)
        if xlsm_path:
            print(f"[csv_exporter] ✓ LLM 自动映射到：{Path(xlsm_path).name}")
        else:
            print("[csv_exporter] ✗ LLM 映射也失败，跳过 loadsheet 生成")
            return None

    # 2. 找类目 ID 字符串
    main_cat, low_cat = find_category_ids(xlsm_path, category_path)

    # 3. 图片 URL
    urls = image_urls or source.image_urls or []

    # 3b. 用 1688 包装信息（第一行）覆盖 draft.attributes 里的尺寸/重量，优先于 LLM 估算
    pkg_list = getattr(source, "packaging_info", None) or []
    if pkg_list:
        pkg = pkg_list[0]
        _pkg_map = {
            "packaged_length":  pkg.get("length_cm", ""),
            "packaged_width":   pkg.get("width_cm", ""),
            "packaged_height":  pkg.get("height_cm", ""),
            "packaged_weight":  pkg.get("weight_g", ""),
        }
        combined = dict(draft.attributes or {})
        for k, v in _pkg_map.items():
            if v and not combined.get(k):   # 只在 LLM 没给值时才覆盖
                combined[k] = str(v)
        draft.attributes = combined

    # 3c. 从 1688 product_attrs 解析重量/尺寸（当 packaging_info 为空时的兜底）
    src_attrs = (getattr(source, "product_attrs", None) or
                 (source.raw_data or {}).get("product_attrs", {}))
    if src_attrs:
        combined = dict(draft.attributes or {})
        # 重量：解析"重量"字段（支持"单218g+带盒385g"格式）
        if not combined.get("packaged_weight") and not combined.get("weight_g"):
            raw_w = str(src_attrs.get("重量", "") or src_attrs.get("净重", "")).strip()
            if raw_w:
                combined["packaged_weight"] = _parse_weight_g(raw_w)
        # 尺寸：解析"外观尺寸"/"产品尺寸"（支持"20×10×5cm"或"20*10*5"格式）
        if not combined.get("packaged_length"):
            raw_dim = str(src_attrs.get("外观尺寸", "") or src_attrs.get("产品尺寸", "")
                          or src_attrs.get("包装尺寸", "")).strip()
            if raw_dim:
                parts = re.findall(r"\d+(?:\.\d+)?", raw_dim)
                if len(parts) >= 3:
                    combined.setdefault("packaged_length", parts[0])
                    combined.setdefault("packaged_width",  parts[1])
                    combined.setdefault("packaged_height", parts[2])
        draft.attributes = combined

    # 4. 构建行数据
    row_values = build_row_values(draft, source, main_cat, low_cat, urls)

    # 4b. 通用 portal 字段透传：把 draft.attributes 里的 portal label 值映射到 xlsm 列
    #     例如 "Screen Size" -> "Attribute.screen_size.value"，适用于所有类目
    row_values = _apply_portal_field_overrides(row_values, draft.attributes or {}, xlsm_path)

    # 4c. 统一清洗数值字段：去掉单位，只保留数字
    for key in list(row_values.keys()):
        v = row_values.get(key, "")
        if not isinstance(v, str):
            continue
        # 打包尺寸 / 产品尺寸 / 视角等，要求纯数字
        if any(
            frag in key
            for frag in [
                "Attribute.merchant_packaged_dimensions.width",
                "Attribute.merchant_packaged_dimensions.length",
                "Attribute.merchant_packaged_dimensions.height",
                "Attribute.product_dimensions.width",
                "Attribute.product_dimensions.length",
                "Attribute.product_dimensions.height",
                "Attribute.view_angle.value",
                "Attribute.input_voltage.value",
                "Attribute.output_voltage.value",
                "Attribute.rated_voltage.value",
            ]
        ):
            row_values[key] = _parse_cm_number(v)

    # 4c. 不再在 loadsheet 层面做任何 AI 补填：
    #     所有字段值都来自：
    #       - 1688 抓取 + 规则推断（build_row_values）
    #       - 预览对话框中用户确认的 portal 字段（_apply_portal_field_overrides）
    #     如果某些必填项仍为空，让平台在上传时报错，再在预览里补充，
    #     避免在 xlsm 里出现用户没看到的“AI 猜测值”。

    # 5. 写文件
    run_dir.mkdir(parents=True, exist_ok=True)
    sku = row_values.get("SKU", "product")
    safe_sku = re.sub(r"[^\w\-]", "_", sku)
    out_name = Path(xlsm_path).stem  # e.g. loadsheet-107
    out_path = run_dir / f"upload_{safe_sku}_{out_name}.xlsm"

    # 构建必填列归一化集合：从 probe_fields 提取必填标签 → 用 _apply_portal_field_overrides
    # 内部相同的归一化规则转换为 xlsm 列的规范化 key，让 write_to_xlsm 可以标橙底色提醒。
    required_portal_norms: set[str] = set()
    _attrs = draft.attributes or {}
    _pf_raw = _attrs.get("_probe_fields") if isinstance(_attrs, dict) else None
    if isinstance(_pf_raw, list):
        for _pf in _pf_raw:
            if not _pf.get("required"):
                continue
            _lbl = str(_pf.get("label", "")).strip()
            if not _lbl:
                continue
            _n = re.sub(r"[/]+", " ", _lbl.lower())
            _n = re.sub(r"[&'\"()]+", "", _n).strip()
            _n = re.sub(r"[\s\-]+", "_", _n)
            required_portal_norms.add(_n)

    try:
        write_to_xlsm(xlsm_path, row_values, out_path, data_start_row=8,
                      required_portal_norms=required_portal_norms or None)
        print(f"[csv_exporter] ✓ 生成：{out_path.name}")
        print(f"    类目：{main_cat} / {low_cat}")
        print(f"    模板：{Path(xlsm_path).name}")
        return out_path
    except Exception as e:
        print(f"[csv_exporter] ✗ 写入失败：{e}")
        return None
