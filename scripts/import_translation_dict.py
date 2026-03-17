#!/usr/bin/env python3
"""
从本地翻译词典 Excel 中导入英文→中文映射，写入 CSV 翻译缓存。

用途：
    - 读取 /Users/wangfugui/Desktop/翻译词典.xlsx
    - 只针对当前 takealot_categories.csv 里实际用到的值
    - 根据列类型填充到缓存：
        * Division
        * Loadsheet/Department
        * Main Category
        * Lowest Category
    - 这样 translate_takealot_categories.py 在运行时，优先使用本地词典，
      只对词典没有覆盖的少数类目调用 LLM。
"""
from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
INPUT_CSV = ROOT / "input" / "takealot_categories.csv"
CACHE_PATH = ROOT / "output" / "takealot_categories_zh_cache.json"
DICT_PATH = Path("/Users/wangfugui/Desktop/翻译词典.xlsx")
EXTRA_JSON_PATH = Path("/Users/wangfugui/Desktop/翻译.txt")


def _strip_id(text: str) -> str:
    """去掉末尾括号里的数字 ID，例如 'Ear Care (16991)' → 'Ear Care'。"""
    return re.sub(r"\s*\(\d+\)\s*$", "", str(text or "")).strip()


def _load_dict() -> Dict[str, str]:
    """
    从 Excel 加载英文→中文词典。

    每一行格式类似：
        'Consumables → 消耗品'
        '- Fitness & Nutrition->Vitamins & Supplements->Bones & Joint Health → 健身与营养->维生素与补充剂->骨骼与关节健康'
    """
    if not DICT_PATH.exists():
        raise SystemExit(f"翻译词典不存在：{DICT_PATH}")

    wb = load_workbook(DICT_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    mapping: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        val = row[0]
        if not isinstance(val, str):
            continue
        s = val.strip()
        if not s or s.startswith("#") or s.startswith("["):
            continue
        if "→" not in s:
            continue
        left, right = s.split("→", 1)
        eng = left.strip().lstrip("-").strip()
        zh = right.strip()
        if not eng or not zh:
            continue
        # 后遇到的翻译覆盖先前的（以词典最新内容为准）
        mapping[eng] = zh

    wb.close()
    print(f"[dict] Excel 词条: {len(mapping)} 条")
    return mapping


def _load_extra_json(mapping: Dict[str, str]) -> Dict[str, str]:
    """
    读取额外的 JSON 词典（/Users/wangfugui/Desktop/翻译.txt），结构为:
        { "English ...": "中文 ...", ... }

    后载入的 JSON 会覆盖 Excel 中的同名键。
    """
    if not EXTRA_JSON_PATH.exists():
        return mapping
    try:
        data = json.loads(EXTRA_JSON_PATH.read_text("utf-8"))
        if isinstance(data, dict):
            for k, v in data.items():
                eng = str(k or "").strip()
                zh = str(v or "").strip()
                if not eng or not zh:
                    continue
                mapping[eng] = zh
            print(f"[dict] JSON 词条: {len(data)} 条（合并后共 {len(mapping)} 条）")
    except Exception as e:
        print(f"[dict] 读取 JSON 词典失败: {e}")
    return mapping


def _load_cache() -> Dict[str, Dict[str, str]]:
    if not CACHE_PATH.exists():
        return {}
    try:
        data = json.loads(CACHE_PATH.read_text("utf-8"))
        if isinstance(data, dict):
            return {
                str(col): {str(k): str(v) for k, v in (mapping or {}).items()}
                for col, mapping in data.items()
                if isinstance(mapping, dict)
            }
    except Exception:
        pass
    return {}


def _save_cache(cache: Dict[str, Dict[str, str]]) -> None:
    CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def _load_csv_values() -> Dict[str, List[str]]:
    if not INPUT_CSV.exists():
        raise SystemExit(f"输入 CSV 不存在：{INPUT_CSV}")

    with INPUT_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))

    if len(rows) < 3:
        raise SystemExit("CSV 行数不足，无法读取表头。")

    header = rows[2]

    def idx_of(name: str) -> int:
        try:
            return header.index(name)
        except ValueError:
            raise SystemExit(f"表头中未找到列名：{name!r}")

    idx_div = idx_of("Division")
    idx_dept = idx_of("Loadsheet/Department")
    idx_main = idx_of("Main Category")
    idx_low = idx_of("Lowest Category")

    data_rows = rows[3:]

    divisions = [r[idx_div] for r in data_rows if len(r) > idx_div]
    departments = [r[idx_dept] for r in data_rows if len(r) > idx_dept]
    mains = [r[idx_main] for r in data_rows if len(r) > idx_main]
    lows = [r[idx_low] for r in data_rows if len(r) > idx_low]

    return {
        "Division": divisions,
        "Loadsheet/Department": departments,
        "Main Category": mains,
        "Lowest Category": lows,
    }


def import_dict() -> None:
    dict_map = _load_dict()
    dict_map = _load_extra_json(dict_map)
    csv_values = _load_csv_values()
    cache = _load_cache()

    for kind, values in csv_values.items():
        col_cache = cache.get(kind, {})
        uniq = sorted({v for v in values if str(v or "").strip()})
        imported = 0
        for v in uniq:
            zh = None

            # 先尝试完整匹配
            if v in dict_map:
                zh = dict_map[v]
            else:
                base = _strip_id(v)
                if base in dict_map:
                    zh = dict_map[base]

            if zh:
                # 词典结果优先：覆盖旧缓存
                if col_cache.get(v) != zh:
                    col_cache[v] = zh
                    imported += 1

        cache[kind] = col_cache
        print(f"[{kind}] 词典可覆盖 {imported} 条，列总唯一值 {len(uniq)} 条。")

    _save_cache(cache)
    print(f"[cache] 已写入：{CACHE_PATH}")


def main() -> None:
    import_dict()


if __name__ == "__main__":
    main()
