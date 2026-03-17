"""
下载 Takealot 所有类目 Loadsheet 并解析成 JSON 模板。

用法：
    python scripts/download_loadsheets.py

需要已登录的 Takealot auth state（.runtime/auth/takealot.json）。
如果还没有，先在主界面点"登录 Takealot 账户"保存登录状态。

输出：
    input/loadsheets/raw/      ← 原始 xlsm 文件
    input/loadsheets/templates/ ← 解析后的 JSON 模板（给 CSV 生成器用）
"""
from __future__ import annotations

import json
import re
import time
from pathlib import Path

import openpyxl
from playwright.sync_api import sync_playwright

# ─── 路径配置（与主项目保持一致）──────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
AUTH_STATE = ROOT / ".runtime" / "auth" / "takealot.json"
RAW_DIR = ROOT / "input" / "loadsheets" / "raw"
TPL_DIR = ROOT / "input" / "loadsheets" / "templates"
BULK_URL = "https://sellers.takealot.com/catalogue/templates"
BROWSER_CHANNEL = "msedge"


# ─── xlsm 解析 ────────────────────────────────────────────────────────────────

def _safe_str(v) -> str:
    return str(v).strip() if v is not None else ""


def parse_loadsheet(xlsm_path: Path) -> dict:
    """
    解析一个 loadsheet xlsm，返回结构化模板。

    Loadsheet 结构：
      Row 1: 机器列名  (Variant.ProductVariant, SKU, ...)
      Row 2: 空 或 说明
      Row 3: 人类可读标签 (Product or Variant, Your own SKU, ...)
      Row 4: 字段说明/备注（如 id_in_brackets）
    """
    wb = openpyxl.load_workbook(str(xlsm_path), read_only=True, keep_vba=False)

    # 优先用 "Loadsheet" sheet，否则取第一个
    sheet_name = "Loadsheet" if "Loadsheet" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))

    machine_row = rows[0] if len(rows) > 0 else []
    label_row   = rows[2] if len(rows) > 2 else []
    note_row    = rows[3] if len(rows) > 3 else []

    columns = []
    for i, key in enumerate(machine_row):
        key_str = _safe_str(key)
        if not key_str:
            continue
        label = _safe_str(label_row[i]) if i < len(label_row) else ""
        note  = _safe_str(note_row[i])  if i < len(note_row)  else ""
        columns.append({
            "index": i + 1,          # 1-based，方便对照 Excel
            "key":   key_str,
            "label": label,
            "note":  note,
        })

    # 从 Lookup sheet 里提取每列的合法值（下拉选项）
    valid_values: dict[str, list[str]] = {}
    if "Lookup" in wb.sheetnames:
        lws = wb["Lookup"]
        lrows = list(lws.iter_rows(min_row=1, max_row=200, values_only=True))
        if lrows:
            # Lookup 第一行是列头（对应 Loadsheet 列名）
            lheader = [_safe_str(v) for v in lrows[0]]
            for col_idx, col_key in enumerate(lheader):
                if not col_key:
                    continue
                vals = []
                for lr in lrows[1:]:
                    v = _safe_str(lr[col_idx]) if col_idx < len(lr) else ""
                    if v:
                        vals.append(v)
                if vals:
                    valid_values[col_key] = vals

    # 标记图片列、变体列、属性列等区域
    image_cols    = [c["key"] for c in columns if re.match(r"Images\.image_url_\d+", c["key"])]
    attr_cols     = [c["key"] for c in columns if c["key"].startswith("Attribute.")]
    variant_cols  = [c["key"] for c in columns if c["key"].startswith("Variant.")]
    required_cols = [c["key"] for c in columns if "id_in_brackets" in c.get("note", "").lower()]

    wb.close()

    return {
        "source_file": xlsm_path.name,
        "sheet_used":  sheet_name,
        "total_columns": len(columns),
        "columns": columns,
        "valid_values": valid_values,
        "regions": {
            "image_columns":   image_cols,
            "attribute_columns": attr_cols,
            "variant_columns": variant_cols,
            "category_id_columns": required_cols,
        },
    }


# ─── Playwright 下载 ──────────────────────────────────────────────────────────

def _safe_filename(text: str) -> str:
    """把部门名变成合法文件名。"""
    t = re.sub(r"[^\w\s-]", "", text.lower())
    t = re.sub(r"[\s]+", "_", t.strip())
    return t[:80]


def download_all_loadsheets(headless: bool = False) -> list[dict]:
    """
    打开 Takealot 批量上传页，展开所有 Division，
    点击每个 Download Loadsheet 按钮，保存文件。

    返回每个下载项的元数据列表。
    """
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    TPL_DIR.mkdir(parents=True, exist_ok=True)

    if not AUTH_STATE.exists():
        raise FileNotFoundError(
            f"找不到 Takealot 登录状态：{AUTH_STATE}\n"
            "请先在主界面点「登录 Takealot 账户」保存登录状态。"
        )

    results = []

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            channel=BROWSER_CHANNEL,
            headless=headless,
            ignore_default_args=["--enable-automation"],
            args=["--disable-blink-features=AutomationControlled", "--disable-infobars"],
        )
        context = browser.new_context(
            storage_state=str(AUTH_STATE),
            viewport={"width": 1440, "height": 900},
            accept_downloads=True,
        )
        page = context.new_page()

        print(f"[→] 打开批量上传页：{BULK_URL}")
        page.goto(BULK_URL, wait_until="domcontentloaded", timeout=60_000)
        page.wait_for_timeout(2000)

        # 展开所有折叠区：先点顶级 Division，再点所有未展开的子级
        print("[→] 展开所有类目分组...")

        def _expand_all_accordions():
            """展开页面上所有 aria-expanded=false 的折叠按钮。"""
            expanded = 0
            collapsed = page.locator("[aria-expanded='false']").all()
            for el in collapsed:
                try:
                    el.scroll_into_view_if_needed(timeout=2000)
                    el.click(timeout=2000)
                    page.wait_for_timeout(300)
                    expanded += 1
                except Exception:
                    pass
            return expanded

        # 多轮展开，直到没有新的折叠项
        for round_i in range(5):
            n = _expand_all_accordions()
            page.wait_for_timeout(600)
            if n == 0:
                break
            print(f"    第{round_i+1}轮展开了 {n} 个折叠区")

        page.wait_for_timeout(1000)

        # 找所有 Download Loadsheet 按钮
        print("[→] 扫描 Download Loadsheet 按钮...")
        btn_locator = page.locator("button:has-text('Download Loadsheet'), a:has-text('Download Loadsheet')")
        total = btn_locator.count()
        print(f"    找到 {total} 个按钮")

        for idx in range(total):
            btn = btn_locator.nth(idx)

            # 找最近的部门名
            dept_name = ""
            try:
                dept_name = btn.evaluate("""
(el) => {
    let cur = el.parentElement;
    for (let i = 0; i < 10 && cur; i++) {
        const h = cur.querySelector('h2,h3,h4,strong,p,[class*="title" i],[class*="name" i],[class*="heading" i]');
        if (h) {
            const t = (h.textContent || '').trim();
            if (t && t.length < 80) return t;
        }
        cur = cur.parentElement;
    }
    return '';
}
""")
            except Exception:
                pass
            dept_name = dept_name.strip() or f"unknown_{idx+1}"
            safe_name = _safe_filename(dept_name)
            print(f"  [{idx+1}/{total}] 下载：{dept_name}")

            try:
                # 滚动到可见区域，确保按钮可点击
                btn.scroll_into_view_if_needed(timeout=5000)
                page.wait_for_timeout(400)

                with page.expect_download(timeout=30_000) as dl_info:
                    btn.click(timeout=8000)
                download = dl_info.value

                orig_name = download.suggested_filename or f"loadsheet_{safe_name}.xlsm"
                dest = RAW_DIR / f"{safe_name}__{orig_name}"
                download.save_as(str(dest))
                print(f"    ✓ 已保存：{dest.name}")

                results.append({
                    "dept": dept_name,
                    "safe_name": safe_name,
                    "file": str(dest),
                    "ok": True,
                })
            except Exception as e:
                print(f"    ✗ 失败：{e}")
                results.append({"dept": dept_name, "ok": False, "error": str(e)})

            time.sleep(1.2)

        context.close()
        browser.close()

    return results


# ─── 解析所有已下载的 xlsm ────────────────────────────────────────────────────

def parse_all_downloaded() -> list[dict]:
    """
    解析 raw/ 目录里所有 xlsm，把模板 JSON 写到 templates/ 目录。
    """
    xlsm_files = sorted(RAW_DIR.glob("*.xlsm"))
    if not xlsm_files:
        print("[!] raw/ 目录没有 xlsm 文件，请先运行下载步骤。")
        return []

    summaries = []
    for xlsm in xlsm_files:
        print(f"[解析] {xlsm.name} ...")
        try:
            tpl = parse_loadsheet(xlsm)
            # JSON 文件名：去掉 xlsm 后缀
            stem = re.sub(r"\.xlsm$", "", xlsm.name, flags=re.IGNORECASE)
            out_path = TPL_DIR / f"{stem}.json"
            out_path.write_text(
                json.dumps(tpl, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            print(f"    ✓ {tpl['total_columns']} 列 → {out_path.name}")
            summaries.append({
                "xlsm": xlsm.name,
                "template": out_path.name,
                "columns": tpl["total_columns"],
                "ok": True,
            })
        except Exception as e:
            print(f"    ✗ 解析失败：{e}")
            summaries.append({"xlsm": xlsm.name, "ok": False, "error": str(e)})

    return summaries


# ─── 主入口 ───────────────────────────────────────────────────────────────────

def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(description="下载并解析 Takealot 所有 Loadsheet")
    parser.add_argument(
        "--parse-only",
        action="store_true",
        help="跳过下载，只解析 raw/ 目录里已有的 xlsm 文件",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="无头模式运行（默认显示浏览器）",
    )
    args = parser.parse_args()

    if not args.parse_only:
        print("=" * 60)
        print("步骤 1/2：下载所有 Loadsheet")
        print("=" * 60)
        dl_results = download_all_loadsheets(headless=args.headless)
        ok_count = sum(1 for r in dl_results if r.get("ok"))
        print(f"\n下载完成：{ok_count}/{len(dl_results)} 成功")
        (RAW_DIR / "_download_log.json").write_text(
            json.dumps(dl_results, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    else:
        print("[跳过下载，直接解析]")

    print("\n" + "=" * 60)
    print("步骤 2/2：解析 xlsm → JSON 模板")
    print("=" * 60)
    parse_results = parse_all_downloaded()
    ok_count = sum(1 for r in parse_results if r.get("ok"))
    print(f"\n解析完成：{ok_count}/{len(parse_results)} 成功")

    # 生成一个汇总索引，方便后续 csv_exporter 查表
    index = {
        r["template"].replace(".json", ""): {
            "template_file": r["template"],
            "columns": r.get("columns", 0),
        }
        for r in parse_results if r.get("ok")
    }
    index_path = TPL_DIR / "_index.json"
    index_path.write_text(json.dumps(index, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n模板索引：{index_path}")
    print("\n所有模板已就绪，下一步可以运行 csv_exporter.py 生成上传文件。")


if __name__ == "__main__":
    main()
