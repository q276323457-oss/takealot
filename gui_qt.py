#!/usr/bin/env python3
"""
Takealot Auto Lister — 桌面控制台
功能：文本 AI（豆包）+ 生图 AI（Gemini）配置 + 1688 链接一键生成 Takealot loadsheet
"""
from __future__ import annotations

import json
import os
import re
import signal
import subprocess
import sys
import threading
import time
import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from PySide6.QtCore import Qt, QTimer, Signal, QObject
from PySide6.QtGui import QColor, QFont, QTextCharFormat, QTextCursor
from PySide6.QtWidgets import (
    QApplication,
    QComboBox,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QListWidget,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QPlainTextEdit,
    QSplitter,
    QVBoxLayout,
    QWidget,
    QInputDialog,
)

if getattr(sys, "frozen", False):
    ROOT = Path(sys.executable).resolve().parent
else:
    ROOT = Path(__file__).resolve().parent


def _default_work_root() -> Path:
    if not getattr(sys, "frozen", False):
        return ROOT
    if sys.platform.startswith("darwin"):
        return Path.home() / "Library" / "Application Support" / "TakealotAutoLister"
    if sys.platform.startswith("win"):
        return Path(os.getenv("APPDATA", str(Path.home()))) / "TakealotAutoLister"
    return Path.home() / ".takealot-autolister"


_work_override = os.getenv("TAKEALOT_APP_HOME", "").strip()
WORK_ROOT = Path(_work_override).expanduser() if _work_override else _default_work_root()
WORK_ROOT.mkdir(parents=True, exist_ok=True)

ENV_FILE = WORK_ROOT / ".env"
if ENV_FILE.exists():
    load_dotenv(ENV_FILE, override=True)
else:
    # 开发环境下仍兼容项目根目录 .env
    load_dotenv(ROOT / ".env", override=False)

RUNS_DIR   = WORK_ROOT / "output" / "runs"
LOG_DIR    = WORK_ROOT / "logs"
CONFIG_FILE = WORK_ROOT / ".runtime" / "ui_config.json"
APP_VERSION = os.getenv("APP_VERSION", "1.0.0").strip() or "1.0.0"
APP_PRODUCT = os.getenv("LICENSE_PRODUCT", "takealot-autolister").strip() or "takealot-autolister"
LICENSE_FILE = WORK_ROOT / ".runtime" / "license.json"
LICENSE_PUBKEY = ROOT / "config" / "license_public.pem"


# ── 信号桥（子线程 → 主线程）────────────────────────────────────────────────

class _Bridge(QObject):
    log_line   = Signal(str, str)   # (text, level: info/ok/warn/err)
    run_done   = Signal(bool)       # success
    btn_state  = Signal(bool, str)  # (enabled, text)
    refresh_runs = Signal()         # 请求主线程刷新历史记录
    update_result = Signal(object)  # updater.UpdateInfo
    update_error = Signal(str)      # 更新检查错误
    # 请求主线程打开预览/编辑对话框（传递 PreviewData + ImageGeneratorSession）
    preview_request = Signal(object, object)


# ── 主窗口 ───────────────────────────────────────────────────────────────────

class MainWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("西安众创南非Takealot自建链接AI工具")
        self.resize(1000, 700)
        self._running = False
        self._bridge = _Bridge()
        self._bridge.log_line.connect(self._append_log)
        self._bridge.run_done.connect(self._on_run_done)
        self._bridge.btn_state.connect(self._set_btn_state)
        self._bridge.refresh_runs.connect(self._refresh_runs)
        self._bridge.update_result.connect(self._on_update_result)
        self._bridge.update_error.connect(self._on_update_error)
        self._bridge.preview_request.connect(self._on_preview_request)
        # threading.Event + result holder for preview_callback synchronisation
        self._preview_event: threading.Event | None = None
        self._preview_result: object = None
        self._license_valid = False
        self._license_message = "未检查"
        # 持有 PreviewDialog 的强引用，防止后台线程触发 GC 时在非主线程析构 Qt 控件
        # 只在主线程的 _on_run_done 里释放
        self._active_preview_dlg: object = None

        self._build_ui()
        self._load_config()
        self._refresh_license_status()
        self._refresh_runs()
        self._tak_status_timer = QTimer(self)
        self._tak_status_timer.setInterval(20_000)
        self._tak_status_timer.timeout.connect(self._refresh_tak_status)
        self._tak_status_timer.start()
        QTimer.singleShot(1200, lambda: self._check_update(silent=True))

    # ── UI ──────────────────────────────────────────────────────────────────

    def _build_ui(self) -> None:
        root = QWidget()
        self.setCentralWidget(root)
        lay = QVBoxLayout(root)
        lay.setContentsMargins(12, 12, 12, 12)
        lay.setSpacing(8)

        # ── 标题 ──
        title_row = QHBoxLayout()
        title = QLabel("🛒  西安众创南非Takealot自建链接AI工具")
        title.setStyleSheet("font-size:18px; font-weight:700; padding:4px 0;")
        title_row.addWidget(title)
        title_row.addStretch(1)
        self.version_label = QLabel(f"v{APP_VERSION}")
        self.version_label.setStyleSheet("color:#666; padding-right:8px;")
        title_row.addWidget(self.version_label)
        btn_check_update = QPushButton("⬇️ 检查更新")
        btn_check_update.setFixedWidth(110)
        btn_check_update.clicked.connect(lambda: self._check_update(silent=False))
        title_row.addWidget(btn_check_update)
        lay.addLayout(title_row)

        contact = QLabel("联系作者微信号：Tkwangfg")
        contact.setStyleSheet("color:#444; font-size:12px; padding:0 0 6px 2px;")
        lay.addWidget(contact)

        # ── 授权状态卡片 ──
        lic_frame = QFrame()
        lic_frame.setFrameShape(QFrame.Shape.StyledPanel)
        lic_lay = QHBoxLayout(lic_frame)
        lic_lay.addWidget(self._section_label("🔐 授权状态："))
        self.license_status_label = QLabel("检查中")
        self.license_status_label.setStyleSheet("font-weight:600; color:#e67e22;")
        lic_lay.addWidget(self.license_status_label)
        lic_lay.addSpacing(12)
        lic_lay.addWidget(QLabel("机器码"))
        self.machine_code_input = QLineEdit()
        self.machine_code_input.setReadOnly(True)
        self.machine_code_input.setMinimumWidth(290)
        lic_lay.addWidget(self.machine_code_input)
        btn_copy_mc = QPushButton("复制机器码")
        btn_copy_mc.clicked.connect(self._copy_machine_code)
        lic_lay.addWidget(btn_copy_mc)
        btn_activate = QPushButton("输入卡密激活")
        btn_activate.setStyleSheet("font-weight:600;")
        btn_activate.clicked.connect(self._activate_license)
        lic_lay.addWidget(btn_activate)
        lic_lay.addStretch(1)
        lay.addWidget(lic_frame)

        # ── 文本/生图配置卡片 ──
        ai_frame = QFrame()
        ai_frame.setFrameShape(QFrame.Shape.StyledPanel)
        ai_grid = QGridLayout(ai_frame)
        ai_grid.setColumnStretch(1, 1)

        ai_grid.addWidget(self._section_label("🤖  AI 配置（文本 + 生图）"), 0, 0, 1, 6)

        ai_grid.addWidget(QLabel("文本 Key（豆包）"), 1, 0)
        self.key_input = QLineEdit()
        self.key_input.setPlaceholderText("b0db4c09-xxxx-xxxx-xxxx-xxxxxxxxxxxx")
        self.key_input.setEchoMode(QLineEdit.EchoMode.Password)
        ai_grid.addWidget(self.key_input, 1, 1)

        btn_eye = QPushButton("👁 显示")
        btn_eye.setFixedWidth(72)
        btn_eye.clicked.connect(self._toggle_key)
        ai_grid.addWidget(btn_eye, 1, 2)

        ai_grid.addWidget(QLabel("  文本模型"), 1, 3)
        self.model_combo = QComboBox()
        self.model_combo.addItems([
            "doubao-seed-2-0-pro-260215",
            "doubao-pro-32k",
            "doubao-lite-32k",
        ])
        self.model_combo.setFixedWidth(210)
        ai_grid.addWidget(self.model_combo, 1, 4)

        btn_save = QPushButton("💾  保存配置")
        btn_save.setFixedWidth(110)
        btn_save.setStyleSheet("font-weight:600;")
        btn_save.clicked.connect(self._save_config)
        ai_grid.addWidget(btn_save, 1, 5)

        ai_grid.addWidget(QLabel("生图 Key（Gemini）"), 2, 0)
        self.gemini_key_input = QLineEdit()
        self.gemini_key_input.setPlaceholderText("sk-xxxx（用于 Gemini 生图）")
        self.gemini_key_input.setEchoMode(QLineEdit.EchoMode.Password)
        ai_grid.addWidget(self.gemini_key_input, 2, 1, 1, 4)

        btn_gemini_eye = QPushButton("👁 显示")
        btn_gemini_eye.setFixedWidth(72)
        btn_gemini_eye.clicked.connect(self._toggle_gemini_key)
        ai_grid.addWidget(btn_gemini_eye, 2, 5)

        lay.addWidget(ai_frame)

        # ── 1688 登录卡片 ──
        login_frame = QFrame()
        login_frame.setFrameShape(QFrame.Shape.StyledPanel)
        login_lay = QHBoxLayout(login_frame)

        login_lay.addWidget(self._section_label("🔐  1688 登录状态："))

        self.login_status_label = QLabel("未知")
        self.login_status_label.setStyleSheet("font-weight:600; color:#e67e22;")
        login_lay.addWidget(self.login_status_label)
        login_lay.addStretch(1)

        btn_login = QPushButton("🌐  登录 1688（弹出浏览器）")
        btn_login.setStyleSheet("font-weight:600;")
        btn_login.clicked.connect(self._login_1688)
        login_lay.addWidget(btn_login)

        btn_slider = QPushButton("🔓  过1688滑块验证")
        btn_slider.setStyleSheet("font-weight:600; color:#e67e22;")
        btn_slider.setToolTip("遇到滑块/验证码时点击，会打开浏览器让你手动过验证，完成后自动保存session")
        btn_slider.clicked.connect(self._solve_slider)
        login_lay.addWidget(btn_slider)

        btn_fix = QPushButton("🔧  修复浏览器占用")
        btn_fix.clicked.connect(self._repair_lock)
        login_lay.addWidget(btn_fix)

        lay.addWidget(login_frame)
        self._refresh_login_status()

        # ── Takealot 登录卡片 ──
        tak_frame = QFrame()
        tak_frame.setFrameShape(QFrame.Shape.StyledPanel)
        tak_lay = QHBoxLayout(tak_frame)

        tak_lay.addWidget(self._section_label("🛒  Takealot 登录状态："))

        self.tak_status_label = QLabel("未知")
        self.tak_status_label.setStyleSheet("font-weight:600; color:#e67e22;")
        tak_lay.addWidget(self.tak_status_label)
        tak_lay.addStretch(1)

        btn_tak_login = QPushButton("🌐  登录 Takealot 卖家后台")
        btn_tak_login.setStyleSheet("font-weight:600;")
        btn_tak_login.clicked.connect(self._login_takealot)
        tak_lay.addWidget(btn_tak_login)
        btn_tak_refresh = QPushButton("🔄  刷新状态")
        btn_tak_refresh.clicked.connect(self._refresh_tak_status)
        tak_lay.addWidget(btn_tak_refresh)

        lay.addWidget(tak_frame)
        self._refresh_tak_status()

        # ── 链接 + 执行卡片 ──
        run_frame = QFrame()
        run_frame.setFrameShape(QFrame.Shape.StyledPanel)
        run_grid = QGridLayout(run_frame)
        run_grid.setColumnStretch(1, 1)

        run_grid.addWidget(self._section_label("🔗  1688 商品链接"), 0, 0, 1, 3)

        run_grid.addWidget(QLabel("链接"), 1, 0)
        self.link_input = QLineEdit()
        self.link_input.setPlaceholderText("https://detail.1688.com/offer/XXXXXXXXX.html")
        run_grid.addWidget(self.link_input, 1, 1)

        self.btn_run = QPushButton("🚀  开始生成")
        self.btn_run.setFixedWidth(120)
        self.btn_run.setStyleSheet("font-weight:700; font-size:14px; padding:6px;")
        self.btn_run.clicked.connect(self._start_run)
        run_grid.addWidget(self.btn_run, 1, 2)

        lay.addWidget(run_frame)

        # ── 主体分栏：左=历史记录，右=日志 ──
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # 左：历史运行列表
        left = QWidget()
        ll = QVBoxLayout(left)
        ll.setContentsMargins(0, 0, 0, 0)
        ll.addWidget(self._section_label("📁  历史记录"))
        self.runs_list = QListWidget()
        self.runs_list.itemDoubleClicked.connect(self._open_run)
        ll.addWidget(self.runs_list)
        btn_open = QPushButton("📂  打开目录")
        btn_open.clicked.connect(self._open_run)
        btn_preview = QPushButton("🔍  预览 & 编辑")
        btn_preview.setStyleSheet("font-weight:600; color:#2980b9;")
        btn_preview.clicked.connect(self._preview_run)
        btn_remember_cat = QPushButton("📌  记住当前类目")
        btn_remember_cat.setToolTip("从本次运行的 source.json + portal_result.json 里记住类目映射，后续相同 1688 类目自动用同一路径")
        btn_remember_cat.clicked.connect(self._remember_category_override)
        btn_refresh = QPushButton("🔄  刷新")
        btn_refresh.clicked.connect(self._refresh_runs)
        btn_regen = QPushButton("📊  重新生成 xlsm")
        btn_regen.setToolTip("从已有的 source.json + draft.json 重新生成 xlsm，无需重新采集")
        btn_regen.clicked.connect(self._regen_xlsm)
        btn_row = QHBoxLayout()
        btn_row.addWidget(btn_open)
        btn_row.addWidget(btn_preview)
        btn_row.addWidget(btn_remember_cat)
        btn_row.addWidget(btn_refresh)
        btn_row.addWidget(btn_regen)
        ll.addLayout(btn_row)
        splitter.addWidget(left)

        # 右：运行日志
        right = QWidget()
        rl = QVBoxLayout(right)
        rl.setContentsMargins(0, 0, 0, 0)
        rl.addWidget(self._section_label("📋  运行日志"))
        self.log_box = QPlainTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setFont(QFont("Menlo" if sys.platform == "darwin" else "Consolas", 11))
        rl.addWidget(self.log_box)

        btn_clear = QPushButton("🗑  清空日志")
        btn_clear.clicked.connect(self.log_box.clear)
        rl.addWidget(btn_clear, alignment=Qt.AlignmentFlag.AlignRight)
        splitter.addWidget(right)

        splitter.setSizes([300, 680])
        lay.addWidget(splitter, 1)

    def _section_label(self, text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setStyleSheet("font-weight:700; font-size:13px; padding:4px 0;")
        return lbl

    # ── 更新检查 ────────────────────────────────────────────────────────────

    def _check_update(self, silent: bool = True) -> None:
        threading.Thread(target=self._do_check_update, args=(silent,), daemon=True).start()

    def _do_check_update(self, silent: bool) -> None:
        try:
            sys.path.insert(0, str(ROOT / "src"))
            from takealot_autolister.updater import check_for_update

            info = check_for_update(APP_VERSION)
            self._bridge.update_result.emit((info, silent))
        except Exception as exc:
            if not silent:
                self._bridge.update_error.emit(str(exc))

    def _open_external_url(self, url: str) -> None:
        if not url:
            return
        try:
            if sys.platform.startswith("win"):
                os.startfile(url)  # type: ignore[attr-defined]
            elif sys.platform.startswith("darwin"):
                subprocess.Popen(["open", url])
            else:
                subprocess.Popen(["xdg-open", url])
        except Exception as e:
            QMessageBox.warning(self, "打开失败", f"无法打开链接：{url}\n{e}")

    def _on_update_result(self, payload: object) -> None:
        try:
            info, silent = payload  # type: ignore[misc]
        except Exception:
            return

        if not getattr(info, "has_update", False):
            if not silent:
                QMessageBox.information(self, "已是最新版本", f"当前版本 v{APP_VERSION} 已是最新。")
            return

        latest = str(getattr(info, "latest_version", "") or "")
        notes = str(getattr(info, "notes", "") or "")
        download_url = str(getattr(info, "download_url", "") or "")
        msg = f"发现新版本：v{latest}\n当前版本：v{APP_VERSION}\n"
        if notes:
            msg += f"\n更新说明：\n{notes[:800]}"
        if not download_url:
            msg += "\n\n未提供当前系统下载链接，请联系管理员。"
            QMessageBox.information(self, "发现新版本", msg)
            return
        msg += f"\n\n是否现在打开下载链接？\n{download_url}"
        btn = QMessageBox.question(self, "发现新版本", msg)
        if btn == QMessageBox.StandardButton.Yes:
            self._open_external_url(download_url)

    def _on_update_error(self, msg: str) -> None:
        QMessageBox.warning(self, "检查更新失败", msg[:500])

    # ── 授权 ────────────────────────────────────────────────────────────────

    def _refresh_license_status(self) -> None:
        try:
            sys.path.insert(0, str(ROOT / "src"))
            from takealot_autolister.licensing import check_local_license
            st = check_local_license(
                license_file=str(LICENSE_FILE),
                public_key_file=str(LICENSE_PUBKEY),
                product=APP_PRODUCT,
            )
        except Exception as e:
            self.license_status_label.setText("⚠️ 检查失败")
            self.license_status_label.setStyleSheet("font-weight:600; color:#e67e22;")
            self.machine_code_input.setText("")
            self._license_valid = False
            self._license_message = str(e)
            return

        self.machine_code_input.setText(st.machine_code)
        self._license_valid = bool(st.valid)
        self._license_message = str(st.message)
        if st.valid:
            info = st.payload or {}
            exp = str(info.get("expires_at", "")).strip()
            card = str(info.get("card_id", "")).strip()
            text = "✅ 已激活"
            if card:
                text += f"（{card}）"
            if exp:
                text += f" 到期 {exp}"
            self.license_status_label.setText(text)
            self.license_status_label.setStyleSheet("font-weight:600; color:#27ae60;")
        else:
            self.license_status_label.setText(f"❌ 未激活：{st.message[:28]}")
            self.license_status_label.setStyleSheet("font-weight:600; color:#e74c3c;")

    def _copy_machine_code(self) -> None:
        txt = self.machine_code_input.text().strip()
        if not txt:
            return
        QApplication.clipboard().setText(txt)
        QMessageBox.information(self, "已复制", "机器码已复制到剪贴板。")

    def _activate_license(self) -> None:
        token, ok = QInputDialog.getMultiLineText(
            self,
            "输入卡密激活",
            "请粘贴绑定本机机器码的卡密（授权码）：",
            "",
        )
        if not ok:
            return
        token = str(token or "").strip()
        if not token:
            QMessageBox.warning(self, "提示", "未输入卡密。")
            return
        try:
            sys.path.insert(0, str(ROOT / "src"))
            from takealot_autolister.licensing import activate_and_save
            st = activate_and_save(
                token=token,
                license_file=str(LICENSE_FILE),
                public_key_file=str(LICENSE_PUBKEY),
                product=APP_PRODUCT,
            )
            self._refresh_license_status()
            QMessageBox.information(self, "激活成功", f"{st.message}\n现在可以正常使用软件。")
        except Exception as e:
            QMessageBox.critical(self, "激活失败", str(e)[:500])

    def _ensure_license(self) -> bool:
        self._refresh_license_status()
        if self._license_valid:
            return True
        QMessageBox.warning(
            self,
            "未授权",
            f"当前未激活，无法开始运行。\n\n原因：{self._license_message}\n\n请先复制机器码并输入卡密激活。",
        )
        return False

    # ── 配置 ────────────────────────────────────────────────────────────────

    def _toggle_key(self) -> None:
        if self.key_input.echoMode() == QLineEdit.EchoMode.Password:
            self.key_input.setEchoMode(QLineEdit.EchoMode.Normal)
        else:
            self.key_input.setEchoMode(QLineEdit.EchoMode.Password)

    def _toggle_gemini_key(self) -> None:
        if self.gemini_key_input.echoMode() == QLineEdit.EchoMode.Password:
            self.gemini_key_input.setEchoMode(QLineEdit.EchoMode.Normal)
        else:
            self.gemini_key_input.setEchoMode(QLineEdit.EchoMode.Password)

    def _load_config(self) -> None:
        # 优先加载 .env 当前值，避免旧 ui_config 覆盖新密钥
        env_doubao_key = os.getenv("DOUBAO_API_KEY", "").strip()
        env_doubao_model = os.getenv("DOUBAO_MODEL", "").strip()
        env_gemini_key = os.getenv("GEMINI_IMAGE_API_KEY", "").strip()

        if env_doubao_key:
            self.key_input.setText(env_doubao_key)
        if env_doubao_model:
            idx = self.model_combo.findText(env_doubao_model)
            if idx >= 0:
                self.model_combo.setCurrentIndex(idx)
        if env_gemini_key:
            self.gemini_key_input.setText(env_gemini_key)

        # 如果 .env 没有，再回退到 UI 配置文件
        if CONFIG_FILE.exists():
            try:
                cfg = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
                if not self.key_input.text().strip():
                    self.key_input.setText(cfg.get("doubao_api_key", ""))
                if self.model_combo.currentIndex() < 0:
                    idx = self.model_combo.findText(cfg.get("doubao_model", ""))
                    if idx >= 0:
                        self.model_combo.setCurrentIndex(idx)
                if not self.gemini_key_input.text().strip():
                    self.gemini_key_input.setText(cfg.get("gemini_image_api_key", ""))
            except Exception:
                pass

    def _save_config(self) -> None:
        key   = self.key_input.text().strip()
        model = self.model_combo.currentText().strip()
        if not key:
            QMessageBox.warning(self, "缺少 Key", "请填写文本模型（豆包）Key！")
            return
        gemini_key = self.gemini_key_input.text().strip()
        CONFIG_FILE.parent.mkdir(parents=True, exist_ok=True)
        CONFIG_FILE.write_text(
            json.dumps({"doubao_api_key": key, "doubao_model": model,
                        "gemini_image_api_key": gemini_key}, indent=2),
            encoding="utf-8",
        )
        self._apply_env(key, model, gemini_key)
        QMessageBox.information(self, "保存成功", "✅  配置已保存，下次启动自动加载。")

    def _apply_env(self, key: str, model: str, gemini_key: str = "") -> None:
        """写入 .env 并注入当前进程环境变量。"""
        env_path = ENV_FILE
        env_path.parent.mkdir(parents=True, exist_ok=True)
        lines: list[str] = []
        if env_path.exists():
            lines = env_path.read_text(encoding="utf-8").splitlines()

        def _set(k: str, v: str) -> None:
            for i, line in enumerate(lines):
                if line.startswith(f"{k}="):
                    lines[i] = f"{k}={v}"
                    return
            lines.append(f"{k}={v}")

        _set("DOUBAO_API_KEY",  key)
        _set("DOUBAO_MODEL",    model)
        _set("DOUBAO_VL_MODEL", model)
        if gemini_key:
            _set("GEMINI_IMAGE_API_KEY", gemini_key)
            os.environ["GEMINI_IMAGE_API_KEY"] = gemini_key
        env_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        os.environ["DOUBAO_API_KEY"]  = key
        os.environ["DOUBAO_MODEL"]    = model
        os.environ["DOUBAO_VL_MODEL"] = model

    # ── 1688 登录 ────────────────────────────────────────────────────────────

    _STATE_1688 = Path(os.getenv("STORAGE_STATE_1688",
                       str(WORK_ROOT / ".runtime" / "auth" / "1688.json")))
    _STATE_TAKEALOT = Path(os.getenv("STORAGE_STATE_TAKEALOT",
                           str(WORK_ROOT / ".runtime" / "auth" / "takealot.json")))

    def _refresh_login_status(self) -> None:
        if self._STATE_1688.exists():
            self.login_status_label.setText("✅  已登录")
            self.login_status_label.setStyleSheet("font-weight:600; color:#27ae60;")
        else:
            self.login_status_label.setText("❌  未登录")
            self.login_status_label.setStyleSheet("font-weight:600; color:#e74c3c;")

    # Takealot 把 auth 存在 localStorage（usr_st_auth），不是 cookie。
    # 这里同时检查 cookie 和 localStorage 两种方式。
    _TAK_NON_AUTH_NAMES = frozenset({"__cf_bm", "_cfuvid", "wfx_unq", "_gid", "_fbp"})
    # localStorage 中表示已登录的 key
    _TAK_AUTH_LS_KEYS   = frozenset({"usr_st_auth", "usr_st_usr", "usr_st_slr"})

    def _has_tak_auth(self, state: dict) -> bool:
        """返回 True 表示 storage_state 中含有 Takealot 有效认证（cookie 或 localStorage）。"""
        # 1) 检查 localStorage（Takealot 主要用这里存 JWT）
        for origin in state.get("origins", []):
            if "sellers.takealot.com" in str(origin.get("origin", "")):
                ls = origin.get("localStorage", [])
                if any(item.get("name") in self._TAK_AUTH_LS_KEYS for item in ls):
                    return True
        # 2) 兼容：检查非追踪类 cookie（旧版或其他来源）
        for c in state.get("cookies", []):
            domain = str(c.get("domain", "")).lower()
            name = str(c.get("name", ""))
            if "takealot" not in domain:
                continue
            if name in self._TAK_NON_AUTH_NAMES or name.startswith("_ga"):
                continue
            return True
        return False

    def _refresh_tak_status(self) -> None:
        if not self._STATE_TAKEALOT.exists():
            self.tak_status_label.setText("❌  未登录")
            self.tak_status_label.setStyleSheet("font-weight:600; color:#e74c3c;")
            return
        try:
            d = json.loads(self._STATE_TAKEALOT.read_text(encoding="utf-8"))
            if not self._has_tak_auth(d):
                self.tak_status_label.setText("❌  未登录（会话已过期）")
                self.tak_status_label.setStyleSheet("font-weight:600; color:#e74c3c;")
                return
        except Exception:
            self.tak_status_label.setText("⚠️  状态文件异常，请重新登录")
            self.tak_status_label.setStyleSheet("font-weight:600; color:#e67e22;")
            return
        self.tak_status_label.setText("✅  已登录")
        self.tak_status_label.setStyleSheet("font-weight:600; color:#27ae60;")

    def _login_1688(self) -> None:
        if self._running:
            QMessageBox.information(self, "提示", "请等待当前任务完成后再登录。")
            return
        self._running = True
        self.btn_run.setEnabled(False)
        threading.Thread(target=self._do_login_1688, daemon=True).start()

    def _do_login_1688(self) -> None:
        try:
            py = ROOT / ".venv" / "bin" / "python"
            self._STATE_1688.parent.mkdir(parents=True, exist_ok=True)
            args = [
                str(py), "-m", "takealot_autolister.login_helper",
                "--url",           "https://detail.1688.com",
                "--state-path",    str(self._STATE_1688),
                "--mode",          "1688",
                "--browser-channel", os.getenv("BROWSER_CHANNEL", "msedge"),
                "--wait-seconds",  "300",
                "--stable-hits",   "6",
                "--verify-url",    "https://detail.1688.com",
            ]
            env = os.environ.copy()
            env["PYTHONPATH"] = "src"
            cp = subprocess.run(args, cwd=str(ROOT), text=True,
                                capture_output=True, env=env)
            if cp.returncode == 0:
                self._log("ok", "✅  1688 登录成功，状态已保存。")
                QTimer.singleShot(0, self._refresh_login_status)
            else:
                err = (cp.stdout + cp.stderr).strip()
                self._log("err", f"❌  1688 登录失败：{err[:300]}")
        except Exception as exc:
            self._log("err", f"❌  登录出错：{exc}")
        finally:
            self._running = False
            self._bridge.btn_state.emit(True, "🚀  开始生成")

    def _login_takealot(self) -> None:
        if self._running:
            QMessageBox.information(self, "提示", "请等待当前任务完成后再登录。")
            return
        self._running = True
        self.btn_run.setEnabled(False)
        self._log("info", "[Takealot] 正在打开卖家后台登录浏览器，请在浏览器中完成登录...")
        threading.Thread(target=self._do_login_takealot, daemon=True).start()

    def _do_login_takealot(self) -> None:
        try:
            py = ROOT / ".venv" / "bin" / "python"
            self._STATE_TAKEALOT.parent.mkdir(parents=True, exist_ok=True)
            args = [
                str(py), "-m", "takealot_autolister.login_helper",
                "--url",           "https://sellers.takealot.com",
                "--state-path",    str(self._STATE_TAKEALOT),
                "--mode",          "takealot",
                "--browser-channel", os.getenv("BROWSER_CHANNEL", "msedge"),
                "--wait-seconds",  "180",
                "--stable-hits",   "4",
                "--verify-url",    "https://sellers.takealot.com",
            ]
            env = os.environ.copy()
            env["PYTHONPATH"] = "src"
            cp = subprocess.run(args, cwd=str(ROOT), text=True,
                                capture_output=True, env=env)
            if cp.returncode == 0:
                self._log("ok", "✅  Takealot 登录成功，状态已保存。")
                QTimer.singleShot(0, self._refresh_tak_status)
            else:
                err = (cp.stdout + cp.stderr).strip()
                self._log("err", f"❌  Takealot 登录失败：{err[:300]}")
        except Exception as exc:
            self._log("err", f"❌  登录出错：{exc}")
        finally:
            self._running = False
            self._bridge.btn_state.emit(True, "🚀  开始生成")

    def _solve_slider(self) -> None:
        """用当前输入的1688链接（或首页）弹出浏览器，让用户手动过滑块/验证码。"""
        if self._running:
            QMessageBox.information(self, "提示", "请等待当前任务完成后再操作。")
            return
        link = self.link_input.text().strip()
        url = link if "1688.com" in link else "https://detail.1688.com"
        self._running = True
        self.btn_run.setEnabled(False)
        self._log("info", f"[过滑块] 正在打开浏览器：{url}")
        self._log("info", "  请在弹出的浏览器中完成验证，完成后会自动保存并关闭。")
        threading.Thread(target=self._do_solve_slider, args=(url,), daemon=True).start()

    def _do_solve_slider(self, url: str) -> None:
        try:
            py = ROOT / ".venv" / "bin" / "python"
            self._STATE_1688.parent.mkdir(parents=True, exist_ok=True)
            args = [
                str(py), "-m", "takealot_autolister.login_helper",
                "--url",             url,
                "--state-path",      str(self._STATE_1688),
                "--mode",            "1688",
                "--browser-channel", os.getenv("BROWSER_CHANNEL", "msedge"),
                "--wait-seconds",    "300",
                "--stable-hits",     "12",
                "--verify-url",      url,
            ]
            env = os.environ.copy()
            env["PYTHONPATH"] = "src"
            cp = subprocess.run(args, cwd=str(ROOT), text=True,
                                capture_output=True, env=env)
            if cp.returncode == 0:
                self._log("ok", "✅  验证完成，session 已保存。现在可以点【开始生成】。")
                QTimer.singleShot(0, self._refresh_login_status)
            else:
                err = (cp.stdout + cp.stderr).strip()
                self._log("warn", f"⚠️  验证未完成或超时：{err[:200]}")
        except Exception as exc:
            self._log("err", f"❌  出错：{exc}")
        finally:
            self._running = False
            self._bridge.btn_state.emit(True, "🚀  开始生成")

    def _repair_lock(self) -> None:
        """结束残留浏览器进程，清理 SingletonLock。"""
        try:
            profile_dir = os.getenv("BROWSER_USER_DATA_DIR", "")
            killed = 0
            if profile_dir:
                cp = subprocess.run(["pgrep", "-f", profile_dir],
                                    capture_output=True, text=True)
                for pid_str in (cp.stdout or "").splitlines():
                    try:
                        pid = int(pid_str.strip())
                        if pid != os.getpid():
                            os.kill(pid, signal.SIGTERM)
                            killed += 1
                    except Exception:
                        pass
                # 清理锁文件
                removed = 0
                for lock in Path(profile_dir).rglob("Singleton*"):
                    try:
                        lock.unlink()
                        removed += 1
                    except Exception:
                        pass
                time.sleep(0.5)
                QMessageBox.information(
                    self, "修复完成",
                    f"已终止相关进程 {killed} 个\n已清理锁文件 {removed} 个\n现在可以重新运行。",
                )
            else:
                QMessageBox.information(self, "无需修复", "未配置 BROWSER_USER_DATA_DIR。")
        except Exception as exc:
            QMessageBox.warning(self, "修复失败", str(exc))

    # ── 运行 ────────────────────────────────────────────────────────────────

    def _start_run(self) -> None:
        if self._running:
            return
        if not self._ensure_license():
            return

        key = self.key_input.text().strip()
        if not key:
            QMessageBox.warning(self, "未配置", "请先填写豆包 API Key 并保存！")
            return

        link = self.link_input.text().strip()
        if not link or "1688.com" not in link:
            QMessageBox.warning(self, "链接无效", "请输入有效的 1688 商品链接！")
            return

        self._apply_env(key, self.model_combo.currentText().strip(),
                        self.gemini_key_input.text().strip())

        self.log_box.clear()
        self._running = True
        self.btn_run.setEnabled(False)
        self.btn_run.setText("⏳  运行中...")

        threading.Thread(target=self._run_pipeline, args=(link,), daemon=True).start()

    def _run_pipeline(self, link: str) -> None:
        import io

        class _StdoutCapture(io.TextIOBase):
            """把 print() 输出实时转发到日志桥。"""
            def __init__(self, bridge: _Bridge, orig):
                self._bridge = bridge
                self._orig   = orig
            def write(self, s: str) -> int:
                s = s.rstrip("\n")
                if s.strip():
                    self._bridge.log_line.emit(s, "info")
                if self._orig:
                    try: self._orig.write(s + "\n")
                    except Exception: pass
                return len(s)
            def flush(self): pass

        orig_stdout = sys.stdout
        sys.stdout  = _StdoutCapture(self._bridge, orig_stdout)

        def _pipeline_log(level: str, msg: str) -> None:
            self._bridge.log_line.emit(msg, level)

        def _preview_callback(preview_data, img_session):
            """后台线程调用：发信号给主线程打开对话框，然后阻塞等待结果。"""
            evt = threading.Event()
            self._preview_event = evt
            self._preview_result = None
            self._bridge.preview_request.emit(preview_data, img_session)
            evt.wait(timeout=600)  # 最多等 10 分钟
            return self._preview_result

        try:
            sys.path.insert(0, str(ROOT / "src"))
            load_dotenv(ENV_FILE if ENV_FILE.exists() else (ROOT / ".env"), override=True)

            from takealot_autolister.pipeline import process_one_link
            from takealot_autolister.rules import load_rules

            self._bridge.log_line.emit(f"🚀 开始处理：{link}", "info")

            rules = load_rules(str(ROOT / "config" / "rules.yaml"))
            result = process_one_link(
                link=link,
                output_dir=RUNS_DIR,
                rules=rules,
                use_llm=True,
                headless=False,
                browser_channel=os.getenv("BROWSER_CHANNEL", "msedge"),
                user_data_dir=os.getenv("BROWSER_USER_DATA_DIR", "").strip() or None,
                storage_state_1688=os.getenv("STORAGE_STATE_1688", "").strip() or None,
                storage_state_takealot=str(self._STATE_TAKEALOT) if self._STATE_TAKEALOT.exists() else None,
                remove_bg=False,
                automate_portal_enabled=False,
                selectors_path=str(ROOT / "config" / "selectors.yaml"),
                portal_mode="draft",
                login_wait_seconds=300,
                browser_profile_directory=os.getenv("BROWSER_PROFILE_DIRECTORY", "Default"),
                log_callback=_pipeline_log,
                preview_callback=_preview_callback,
            )

            r = result.__dict__
            action = r.get("action", "")
            ok = r.get("ok", False)
            if r.get("run_dir"):
                self._log("info", f"📁 输出目录：{r['run_dir']}")

            if action == "need_verify_1688":
                self._log("err", "❌  1688 滑块验证超时（等待 5 分钟未完成）。")
                self._log("warn", "   → 下次遇到滑块时，请直接在弹出的浏览器窗口中拖动滑块完成验证，程序会自动继续。")
                self._log("warn", "   → 若浏览器已关闭，请点击「过1688滑块验证」后重新运行。")
                self._bridge.run_done.emit(False)
            elif action in ("need_login_1688",):
                self._log("err", "❌  1688 需要登录，请点击「登录 1688」按钮后重试。")
                self._bridge.run_done.emit(False)
            elif action == "need_retry_1688":
                self._log("warn", f"⚠️  1688 抓取失败（可重试）：{r.get('message', '')}")
                self._bridge.run_done.emit(False)
            elif action == "source_capture_failed":
                self._log("err", f"❌  抓取失败：{r.get('message', '')}")
                self._bridge.run_done.emit(False)
            elif action == "preview_cancelled":
                self._log("warn", "⚠️  您在预览对话框中点了取消，流程已终止。")
                self._bridge.run_done.emit(False)
            elif ok:
                self._log("ok", "")
                self._log("ok", "🎉  完成！xlsm 已生成，请在历史记录中双击打开目录。")
                self._bridge.run_done.emit(True)
            else:
                self._log("warn", f"⚠️  完成但有问题（action={action}）：{r.get('message', '')}")
                self._bridge.run_done.emit(False)

        except Exception as exc:
            import traceback
            self._log("err", f"❌  失败：{exc}")
            self._log("err", traceback.format_exc())
            self._bridge.run_done.emit(False)
        finally:
            sys.stdout = orig_stdout

    def _log(self, level: str, text: str) -> None:
        self._bridge.log_line.emit(text, level)

    def _append_log(self, text: str, level: str) -> None:
        colors = {
            "ok":   "#27ae60",
            "warn": "#e67e22",
            "err":  "#e74c3c",
            "info": "#2980b9",
        }
        color = colors.get(level, "")
        cursor = self.log_box.textCursor()
        cursor.movePosition(QTextCursor.MoveOperation.End)
        fmt = QTextCharFormat()
        if color:
            fmt.setForeground(QColor(color))
        cursor.insertText(text + "\n", fmt)
        self.log_box.setTextCursor(cursor)
        self.log_box.ensureCursorVisible()

    def _set_btn_state(self, enabled: bool, text: str) -> None:
        self.btn_run.setEnabled(enabled)
        if text:
            self.btn_run.setText(text)

    def _on_run_done(self, _success: bool) -> None:
        self._running = False
        self.btn_run.setEnabled(True)
        self.btn_run.setText("🚀  开始生成")
        self._refresh_runs()
        self._active_preview_dlg = None

    def _on_preview_request(self, preview_data: object, img_session: object) -> None:
        """主线程槽：后台流水线请求打开预览/编辑对话框。"""
        dlg = None
        try:
            from takealot_autolister.preview_dialog import PreviewDialog
            from PySide6.QtWidgets import QDialog
            dlg = PreviewDialog(preview_data, img_session, parent=self)
            # 保存强引用：防止 dlg 离开局部作用域后被后台线程的 GC 在非主线程析构
            self._active_preview_dlg = dlg
            dlg.raise_()
            dlg.activateWindow()
            if dlg.exec() == QDialog.DialogCode.Accepted:
                # 深拷贝成纯 Python 数据，避免携带 Qt wrapper 到后台线程
                self._preview_result = copy.deepcopy(dlg.get_result())
            else:
                self._preview_result = None
        except Exception as _e:
            import traceback
            err_msg = traceback.format_exc()
            print(f"[error] 预览对话框打开失败：{_e}\n{err_msg}")
            QMessageBox.critical(self, "预览对话框错误",
                                 f"打开预览对话框时出错：\n{_e}\n\n详情已打印到日志。")
            self._preview_result = None
        finally:
            # 显式在主线程销毁对话框，避免 Qt 对象在后台线程被 GC 触发析构而崩溃
            try:
                if dlg is not None:
                    dlg.deleteLater()
            except Exception:
                pass
            self._active_preview_dlg = None
            # 无论成功/失败都必须唤醒后台线程，否则流水线永久阻塞
            if self._preview_event:
                self._preview_event.set()

    # ── 历史记录 ─────────────────────────────────────────────────────────────

    def _refresh_runs(self) -> None:
        self.runs_list.clear()
        if not RUNS_DIR.exists():
            return
        dirs = sorted(
            [p for p in RUNS_DIR.iterdir() if p.is_dir()],
            reverse=True,
        )[:40]
        for d in dirs:
            tag = ""
            xlsm = list(d.glob("*.xlsm"))
            if xlsm:
                tag = "✅ xlsm"
            elif (d / "draft.json").exists():
                tag = "📄 draft"
            self.runs_list.addItem(f"{d.name}  {tag}")

    def _open_run(self) -> None:
        item = self.runs_list.currentItem()
        if not item:
            return
        name = item.text().split("  ")[0]
        target = RUNS_DIR / name
        if target.exists():
            subprocess.Popen(["open", str(target)])

    def _preview_run(self) -> None:
        item = self.runs_list.currentItem()
        if not item:
            QMessageBox.information(self, "提示", "请先在历史记录中选择一个记录。")
            return
        name = item.text().split("  ")[0]
        target = RUNS_DIR / name
        if not target.exists():
            QMessageBox.warning(self, "目录不存在", f"找不到目录：{target}")
            return

        draft_file  = target / "draft.json"
        source_file = target / "source.json"
        if not draft_file.exists():
            QMessageBox.warning(self, "数据不完整", "找不到 draft.json，无法打开预览。")
            return

        try:
            sys.path.insert(0, str(ROOT / "src"))
            load_dotenv(ENV_FILE if ENV_FILE.exists() else (ROOT / ".env"), override=True)

            from takealot_autolister.preview_dialog import PreviewData, PreviewDialog
            from takealot_autolister.image_generator import ImageGeneratorSession

            draft_data  = json.loads(draft_file.read_text(encoding="utf-8"))
            source_data = json.loads(source_file.read_text(encoding="utf-8")) if source_file.exists() else {}
        except Exception as e:
            QMessageBox.critical(self, "导入/读取失败", str(e))
            return

        attrs = draft_data.get("attributes") or {}
        portal_fields = attrs.get("_probe_fields") or []
        category_path = attrs.get("_category_path") or source_data.get("category_path") or []
        image_urls    = source_data.get("image_urls") or []
        field_values: dict = {
            k: str(v) for k, v in attrs.items()
            if not str(k).startswith("_") and (" " in str(k) or (str(k) and str(k)[0].isupper()))
        }

        preview_data = PreviewData(
            title=draft_data.get("title", ""),
            subtitle=draft_data.get("subtitle", ""),
            source_image_urls=image_urls,
            portal_fields=portal_fields,
            field_values=field_values,
            category_path=category_path,
            product_info=source_data,
            run_dir=str(target),
        )
        img_session = ImageGeneratorSession(
            source_urls=image_urls,
            product_title=draft_data.get("title", ""),
        )

        from PySide6.QtWidgets import QDialog
        dlg = PreviewDialog(preview_data, img_session, parent=self)
        self._active_preview_dlg = dlg
        try:
            if dlg.exec() != QDialog.DialogCode.Accepted:
                return
            result = copy.deepcopy(dlg.get_result())
            if not result:
                return
        finally:
            try:
                dlg.deleteLater()
            except Exception:
                pass
            self._active_preview_dlg = None

        # 把预览结果写回 draft，然后重新生成 xlsm
        if self._running:
            QMessageBox.information(self, "提示", "请等待当前任务完成后再操作。")
            return
        self._running = True
        self._log("info", f"📊 预览确认，重新生成 xlsm：{name}")
        threading.Thread(target=self._do_preview_regen, args=(target, draft_data, source_data, result), daemon=True).start()

    def _do_preview_regen(self, run_dir: Path, draft_data: dict, source_data: dict, result: Any) -> None:
        """预览确认后：写回字段 + 上传 OSS + 生成 xlsm。"""
        try:
            sys.path.insert(0, str(ROOT / "src"))
            load_dotenv(ENV_FILE if ENV_FILE.exists() else (ROOT / ".env"), override=True)

            from takealot_autolister.types import ProductSource, ListingDraft
            from takealot_autolister.csv_exporter import generate_loadsheet
            from takealot_autolister.oss_uploader import upload_bytes_list

            draft = ListingDraft(**{k: v for k, v in draft_data.items()
                                    if k in ListingDraft.__dataclass_fields__})
            source = ProductSource(**{k: v for k, v in source_data.items()
                                      if k in ProductSource.__dataclass_fields__}) if source_data else ProductSource(source_url="", title="")

            if result.title:
                draft.title = result.title
            if result.subtitle:
                draft.subtitle = result.subtitle
            for kf_label in ("Key Selling Features", "key_features", "Key Features"):
                kf_val = (result.field_values or {}).get(kf_label, "").strip()
                if kf_val:
                    draft.key_features = kf_val
                    break
            combined = dict(draft.attributes or {})
            combined.update(result.field_values or {})
            new_cat_path = [str(x).strip() for x in (getattr(result, "category_path", []) or []) if str(x).strip()]
            if new_cat_path:
                combined["_category_path"] = new_cat_path
            new_probe_fields = getattr(result, "portal_fields", None) or []
            if isinstance(new_probe_fields, list) and new_probe_fields:
                combined["_probe_fields"] = new_probe_fields
            draft.attributes = combined

            # 上传 OSS
            oss_urls: list[str] = []
            if result.selected_image_bytes:
                self._log("info", f"  ☁️  上传 {len(result.selected_image_bytes)} 张图到 OSS...")
                oss_urls = upload_bytes_list(result.selected_image_bytes, stem="ai_product")
                if oss_urls:
                    self._log("ok", f"  ✅ 上传完成：{len(oss_urls)} 个 URL")

            xlsm_path = generate_loadsheet(draft, source, run_dir, image_urls=oss_urls or None)
            if xlsm_path:
                self._log("ok", f"  ✅ xlsm 已生成：{Path(xlsm_path).name}")
                self._bridge.refresh_runs.emit()
            else:
                self._log("warn", "  ⚠️  xlsm 生成失败，请检查类目路径")
        except Exception as e:
            self._log("error", f"  ❌ 生成失败：{e}")
        finally:
            self._running = False

    def _regen_xlsm(self) -> None:
        item = self.runs_list.currentItem()
        if not item:
            QMessageBox.information(self, "提示", "请先在历史记录中选择一个记录。")
            return
        name = item.text().split("  ")[0]
        target = RUNS_DIR / name
        if not target.exists():
            QMessageBox.warning(self, "目录不存在", f"找不到目录：{target}")
            return
        source_file = target / "source.json"
        draft_file = target / "draft.json"
        if not source_file.exists() or not draft_file.exists():
            QMessageBox.warning(self, "数据不完整", "找不到 source.json 或 draft.json，无法重新生成。")
            return
        if self._running:
            QMessageBox.information(self, "提示", "请等待当前任务完成后再操作。")
            return
        self._running = True
        self._log("info", f"📊 重新生成 xlsm：{name}")
        threading.Thread(target=self._do_regen_xlsm, args=(target,), daemon=True).start()

    # ── 类目记忆 ────────────────────────────────────────────────────────────────

    def _remember_category_override(self) -> None:
        """将当前 run 的 1688 类目 → Takealot 类目 映射写入 category_overrides.yaml。"""
        item = self.runs_list.currentItem()
        if not item:
            QMessageBox.information(self, "提示", "请先在历史记录中选择一个记录。")
            return
        name = item.text().split("  ")[0]
        run_dir = RUNS_DIR / name
        if not run_dir.exists():
            QMessageBox.warning(self, "目录不存在", f"找不到目录：{run_dir}")
            return

        source_file = run_dir / "source.json"
        portal_result_file = run_dir / "portal_result.json"
        if not source_file.exists():
            QMessageBox.warning(self, "数据不完整", "找不到 source.json，无法读取 1688 类目信息。")
            return
        if not portal_result_file.exists():
            QMessageBox.warning(self, "数据不完整", "找不到 portal_result.json，此次运行尚未执行后台类目匹配。")
            return

        try:
            source_data = json.loads(source_file.read_text(encoding="utf-8"))
            portal_result = json.loads(portal_result_file.read_text(encoding="utf-8"))
        except Exception as e:
            QMessageBox.critical(self, "读取失败", f"读取 JSON 出错：{e}")
            return

        src_cat_path = [str(x).strip() for x in (source_data.get("category_path") or []) if str(x).strip()]
        src_title = str(source_data.get("title", "")).strip()
        if not src_cat_path and not src_title:
            QMessageBox.warning(self, "数据不完整", "source.json 中没有有效的 1688 类目路径或标题，无法记忆。")
            return

        takealot_path = self._derive_selected_takealot_path(portal_result)
        if not takealot_path:
            options = self._build_takealot_path_options(portal_result)
            if not options:
                QMessageBox.warning(
                    self,
                    "没有可用类目候选",
                    "portal_result.json 中没有找到可用的类目路径，无法记忆类目。",
                )
                return
            labels = [lab for lab, _ in options]
            # 自动识别失败时，回退到人工选择
            choice, ok = QInputDialog.getItem(
                self,
                "选择要记住的 Takealot 类目",
                "自动识别当前类目失败，请手动选一条要记住的路径：",
                labels,
                0,
                False,
            )
            if not ok or not choice:
                return
            idx = labels.index(choice)
            takealot_path = options[idx][1]

        # 读写 overrides 文件
        overrides_path = ROOT / "input" / "category_overrides.yaml"
        overrides_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            if overrides_path.exists():
                import yaml  # type: ignore
                data = yaml.safe_load(overrides_path.read_text(encoding="utf-8"))
                if not isinstance(data, list):
                    data = []
            else:
                data = []
        except Exception:
            data = []

        # 简单防重：同一个 source_category_path 已存在就替换
        new_item = {
            "source_category_path": src_cat_path,
            "keywords": [src_title, *src_cat_path],
            "takealot_path": takealot_path,
        }
        merged: list[dict[str, Any]] = []
        replaced = False
        for it in data:
            if (
                isinstance(it, dict)
                and [str(x).strip() for x in it.get("source_category_path", []) if str(x).strip()] == src_cat_path
            ):
                merged.append(new_item)
                replaced = True
            else:
                merged.append(it)
        if not replaced:
            merged.append(new_item)

        try:
            import yaml  # type: ignore
            overrides_path.write_text(yaml.safe_dump(merged, allow_unicode=True, sort_keys=False), encoding="utf-8")
        except Exception as e:
            QMessageBox.critical(self, "写入失败", f"写入 category_overrides.yaml 出错：{e}")
            return

        self._log(
            "ok",
            "📌 已记住类目映射：\n"
            f"  1688 类目: {' > '.join(src_cat_path) or '(无)'}\n"
            f"  Takealot: {' > '.join(takealot_path)}",
        )
        QMessageBox.information(
            self,
            "已记住类目",
            "✅ 已将本次 1688 类目 → Takealot 类目映射写入 input/category_overrides.yaml。\n"
            "后续同一 1688 类目将优先使用该映射自动匹配。",
        )

    def _norm_cat_text(self, text: str) -> str:
        t = str(text or "").lower()
        t = t.replace("›", " ").replace(">", " ").replace("->", " ")
        t = re.sub(r"[^0-9a-z\u4e00-\u9fff]+", " ", t)
        return " ".join(t.split())

    def _path_in_selected_text(self, selected_text: str, path: list[str]) -> bool:
        if not selected_text or not path:
            return False
        s = self._norm_cat_text(selected_text)
        pos = 0
        for seg in path:
            n = self._norm_cat_text(seg)
            if not n:
                continue
            idx = s.find(n, pos)
            if idx < 0:
                return False
            pos = idx + len(n)
        return True

    def _build_takealot_path_options(self, portal_result: dict[str, Any]) -> list[tuple[str, list[str]]]:
        cat_match = portal_result.get("category_match") or {}
        candidates = cat_match.get("top_candidates") or []
        used_path = portal_result.get("category_path_used") or []
        options: list[tuple[str, list[str]]] = []
        if isinstance(candidates, list):
            for c in candidates:
                if not isinstance(c, dict):
                    continue
                path = [str(x).strip() for x in (c.get("path") or []) if str(x).strip()]
                if not path:
                    continue
                label = " / ".join(
                    [
                        str(c.get("division", "")).strip(),
                        str(c.get("department", "")).strip(),
                        str(c.get("main", "")).strip(),
                        str(c.get("lowest", "")).strip(),
                    ]
                ).strip(" /")
                if not label:
                    label = " > ".join(path)
                options.append((label, path))
        if used_path:
            path = [str(x).strip() for x in used_path if str(x).strip()]
            if path and all(path != p for _, p in options):
                options.insert(0, ("(当前使用路径) " + " > ".join(path), path))
        return options

    def _derive_selected_takealot_path(self, portal_result: dict[str, Any]) -> list[str]:
        selected_text = str(portal_result.get("selected_category", "") or "")
        options = self._build_takealot_path_options(portal_result)
        if not selected_text or not options:
            return []
        # 优先找"在 selected_category 文本中按顺序出现"的路径，分数高者优先
        scored: list[tuple[int, int, list[str]]] = []
        for _, path in options:
            if not self._path_in_selected_text(selected_text, path):
                continue
            # 分数：路径段数越多越优先；总字符越长越优先（更具体）
            scored.append((len(path), sum(len(x) for x in path), path))
        if scored:
            scored.sort(reverse=True)
            return scored[0][2]
        # 回退：直接用当前使用路径
        used_path = [str(x).strip() for x in (portal_result.get("category_path_used") or []) if str(x).strip()]
        return used_path

    def _do_regen_xlsm(self, run_dir: Path) -> None:
        try:
            sys.path.insert(0, str(ROOT / "src"))
            load_dotenv(ENV_FILE if ENV_FILE.exists() else (ROOT / ".env"), override=True)

            from takealot_autolister.types import ProductSource, ListingDraft
            from takealot_autolister.csv_exporter import generate_loadsheet
            from takealot_autolister.rules import load_rules

            # 读取 source.json
            source_data = json.loads((run_dir / "source.json").read_text(encoding="utf-8"))
            source = ProductSource(**{k: v for k, v in source_data.items()
                                      if k in ProductSource.__dataclass_fields__})

            # 读取 draft.json
            draft_data = json.loads((run_dir / "draft.json").read_text(encoding="utf-8"))
            draft = ListingDraft(**{k: v for k, v in draft_data.items()
                                    if k in ListingDraft.__dataclass_fields__})

            # 提取已有 xlsm 中的图片 URL（如果有的话）
            image_urls: list[str] = []
            existing_xlsm = list(run_dir.glob("*.xlsm"))
            if existing_xlsm:
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(existing_xlsm[0], keep_vba=True, data_only=True)
                    ws = wb.active
                    # 找表头行（通常在第1-3行）
                    header_row = None
                    for r in range(1, 6):
                        for c in range(1, ws.max_column + 1):
                            v = ws.cell(row=r, column=c).value
                            if v and "image_url_1" in str(v).lower():
                                header_row = r
                                break
                        if header_row:
                            break
                    if header_row:
                        # 找 image_url_1 到 image_url_20 的列
                        col_map = {}
                        for c in range(1, ws.max_column + 1):
                            v = ws.cell(row=header_row, column=c).value
                            if v and "image_url_" in str(v).lower():
                                col_map[str(v).strip()] = c
                        # 读第一数据行
                        data_row = header_row + 1
                        for i in range(1, 21):
                            key = "Images.image_url_1" if i == 1 else f"Images.image_url_{i}"
                            col = col_map.get(key)
                            if col:
                                val = ws.cell(row=data_row, column=col).value
                                if val and str(val).startswith("http"):
                                    image_urls.append(str(val).strip())
                    wb.close()
                except Exception as e:
                    self._log("warn", f"  ⚠️  读取旧 xlsm 图片 URL 失败：{e}，将使用 source 原图")

            # 如果 xlsm 中没有图，退回 source.image_urls
            if not image_urls:
                image_urls = source.image_urls or []

            self._log("info", f"  📷 图片 URL：{len(image_urls)} 张")

            xlsm_path = generate_loadsheet(draft, source, run_dir, image_urls=image_urls or None)
            if xlsm_path:
                self._log("ok", f"  ✅ xlsm 已生成：{Path(xlsm_path).name}")
            else:
                self._log("warn", "  ⚠️  xlsm 生成失败，请检查类目路径")
        except Exception as exc:
            import traceback
            self._log("err", f"❌  重新生成失败：{exc}")
            self._log("err", traceback.format_exc())
        finally:
            self._running = False
            self._bridge.btn_state.emit(True, "🚀  开始生成")
            self._bridge.refresh_runs.emit()


# ── 入口 ─────────────────────────────────────────────────────────────────────

def main() -> None:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    w = MainWindow()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
